import io
from django.utils import timezone
from django.http import HttpResponse
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.pagination import PageNumberPagination
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.exceptions import ValidationError
from django.db.models import Q, Count
from .models import PDLBooking, Exam, ExamBooking, ExamResult
from .serializers import PDLBookingSerializer, ExamSerializer, ExamBookingSerializer, ExamResultSerializer
from students.lifecycle import (
    has_min_payment, get_course_pdl, is_pdl_expired,
    activate_course_on_pdl_approval,
    sync_student_status, apply_exam_result,
)


class StandardPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 200

    def get_paginated_response(self, data):
        return Response({
            "count": self.page.paginator.count,
            "total_pages": self.page.paginator.num_pages,
            "results": data,
        })


class PDLBookingViewSet(viewsets.ModelViewSet):
    serializer_class = PDLBookingSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        qs = PDLBooking.objects.select_related(
            "student", "student_course", "booked_by", "approved_by"
        ).all()
        student_id = self.request.query_params.get("student_id")
        student_course_id = self.request.query_params.get("student_course_id")
        booking_status = self.request.query_params.get("status")
        if student_id:
            qs = qs.filter(student_id=student_id)
        if student_course_id:
            qs = qs.filter(student_course_id=student_course_id)
        if booking_status:
            qs = qs.filter(status=booking_status)
        return qs

    def perform_create(self, serializer):
        student_course = serializer.validated_data.get("student_course")
        if not student_course:
            raise ValidationError({"student_course": "This field is required."})

        if student_course.course.is_refresher_course:
            raise ValidationError({"student_course": "Refresher courses do not require PDL booking."})

        # Min payment check is per-course
        if not has_min_payment(student_course):
            raise ValidationError(
                {"student_course": "Course must have at least 10% payment before adding PDL."}
            )

        # Must be in pending_pdl status
        if student_course.status not in ("pending_pdl", "onboarded"):
            raise ValidationError(
                {"student_course": f"Course is not ready for PDL entry (status: {student_course.status})."}
            )

        # Block if ANY pending or active (approved non-expired) PDL exists
        from datetime import timedelta
        existing = student_course.pdl_bookings.filter(
            Q(status="pending") | Q(status="approved")
        ).exclude(
            status="approved",
            approved_at__lt=timezone.now() - timedelta(days=120)
        )
        if existing.exists():
            raise ValidationError(
                {"student_course": "This course already has an active PDL."}
            )

        # Validate required PDL document fields
        reference_number = serializer.validated_data.get("reference_number", "").strip()
        issued_date = serializer.validated_data.get("issued_date")
        if not reference_number:
            raise ValidationError({"reference_number": "PDL reference number is required."})
        if not issued_date:
            raise ValidationError({"issued_date": "PDL issued date is required."})

        student = student_course.student
        # Save as approved immediately — no admin approval step required
        instance = serializer.save(
            booked_by=self.request.user,
            student=student,
            status="approved",
            approved_by=self.request.user,
            approved_at=timezone.now(),
        )
        # Activate the course immediately
        activate_course_on_pdl_approval(student_course)



class ExamViewSet(viewsets.ModelViewSet):
    serializer_class = ExamSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        params = self.request.query_params
        user = self.request.user
        # Scope booking_count to the user's branch if they are a branch user
        if user.role == "branch_user" and user.branch:
            count_filter = Count("bookings", filter=Q(bookings__student__branch=user.branch))
        else:
            count_filter = Count("bookings")
        qs = Exam.objects.annotate(booking_count=count_filter).order_by("exam_date")
        exam_status = params.get("status")
        if exam_status:
            qs = qs.filter(status=exam_status)
        month = params.get("month")  # expects "YYYY-MM"
        if month:
            try:
                year, m = month.split("-")
                qs = qs.filter(exam_date__year=int(year), exam_date__month=int(m))
            except (ValueError, AttributeError):
                pass
        return qs

    def perform_create(self, serializer):
        serializer.save(created_by=self.request.user)

    @action(detail=True, methods=["get"])
    def export(self, request, pk=None):
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

        exam = self.get_object()
        bookings = ExamBooking.objects.select_related(
            "student__branch", "student_course__course",
            "booked_by", "approved_by",
        ).prefetch_related("result").filter(exam=exam, status="confirmed")
        if request.user.role == "branch_user" and request.user.branch:
            bookings = bookings.filter(student__branch=request.user.branch)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Exam Results"

        headers = [
            "Student Name", "Phone", "Admission No", "ID Number", "Branch", "Course",
            "StudentCourse Status", "Booking Status", "Result",
            "Retake", "Booking Date", "Approved By", "Approved At",
        ]
        header_fill = PatternFill("solid", fgColor="111827")
        header_font = Font(bold=True, color="FFFFFF", size=10)
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for b in bookings:
            try:
                result_val = b.result.result.upper()
            except Exception:
                result_val = "PENDING"

            sc = b.student_course
            is_retake = sc.status == "retake_booked" if sc else False

            ws.append([
                b.student.full_name if b.student else "",
                b.student.phone if b.student else "",
                b.student.admission_number if b.student else "",
                b.student.id_number if b.student else "",
                b.student.branch.name if b.student and b.student.branch else "",
                sc.course.class_name if sc and sc.course else "",
                sc.status if sc else "",
                b.status.capitalize(),
                result_val,
                "Yes" if is_retake else "No",
                b.created_at.strftime("%Y-%m-%d") if b.created_at else "",
                b.approved_by.get_full_name() if b.approved_by else "",
                b.approved_at.strftime("%Y-%m-%d %H:%M") if b.approved_at else "",
            ])

        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max(
                len(str(col[0].value or "")),
                max((len(str(c.value or "")) for c in col[1:]), default=0),
            ) + 4

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        safe_name = exam.exam_name.replace(" ", "_")
        filename = f"{safe_name}_{exam.exam_date}.xlsx"
        response = HttpResponse(
            buffer.read(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response

    @action(detail=False, methods=["get"])
    def summary(self, request):
        """Lightweight pass/fail/pending counts for the selected month — no booking objects loaded."""
        params = request.query_params
        qs = ExamBooking.objects.filter(exam__isnull=False)

        if request.user.role == "branch_user" and request.user.branch:
            qs = qs.filter(student__branch=request.user.branch)
        elif request.user.role == "super_admin":
            branch_id = params.get("branch_id")
            if branch_id:
                qs = qs.filter(student__branch_id=branch_id)

        month = params.get("month")
        if month:
            try:
                year, m = month.split("-")
                qs = qs.filter(exam__exam_date__year=int(year), exam__exam_date__month=int(m))
            except (ValueError, AttributeError):
                pass

        exam_id = params.get("exam_id")
        if exam_id:
            qs = qs.filter(exam_id=exam_id)

        counts = qs.aggregate(
            passed=Count("id", filter=Q(result__result="pass")),
            failed=Count("id", filter=Q(result__result="fail")),
            pending=Count("id", filter=Q(result__isnull=True)),
        )
        return Response(counts)

    @action(detail=True, methods=["post"])
    def close_exam(self, request, pk=None):
        exam = self.get_object()
        exam.status = "closed"
        exam.save()
        return Response({"message": "Exam closed"})


class ExamBookingViewSet(viewsets.ModelViewSet):
    serializer_class = ExamBookingSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = None

    def get_queryset(self):
        user = self.request.user
        params = self.request.query_params
        qs = ExamBooking.objects.select_related(
            "student_course__student__branch",
            "student_course__course",
            "exam", "booked_by",
        ).prefetch_related("result").all()

        if user.role == "branch_user" and user.branch:
            qs = qs.filter(student__branch=user.branch)
        elif user.role == "super_admin":
            branch_id = params.get("branch_id")
            if branch_id:
                qs = qs.filter(student__branch_id=branch_id)

        student_course_id = params.get("student_course_id")
        student_id = params.get("student_id")
        exam_id = params.get("exam_id")
        if student_course_id:
            qs = qs.filter(student_course_id=student_course_id)
        if student_id:
            qs = qs.filter(student__id=student_id)
        if exam_id:
            # Enable pagination only for exam card fetches
            self.pagination_class = StandardPagination
            qs = qs.filter(exam_id=exam_id)
        else:
            booking_status = params.get("status")
            if booking_status:
                qs = qs.filter(status=booking_status)
            elif not params.get("student_course_id") and not params.get("student_id"):
                qs = qs.filter(exam__isnull=False)
                # Month filter for the upfront page-level stats fetch
                month = params.get("month")
                if month:
                    try:
                        year, m = month.split("-")
                        qs = qs.filter(exam__exam_date__year=int(year), exam__exam_date__month=int(m))
                    except (ValueError, AttributeError):
                        pass
            else:
                qs = qs.filter(status="pending")

        result = params.get("result")
        if result:
            if result == "pending":
                qs = qs.filter(result__isnull=True)
            else:
                qs = qs.filter(result__result=result)

        search = params.get("search")
        if search:
            qs = qs.filter(
                Q(student__full_name__icontains=search)
                | Q(student__admission_number__icontains=search)
                | Q(student__phone__icontains=search)
                | Q(student__id_number__icontains=search)
            )
        return qs

    def create(self, request, *args, **kwargs):
        serializer = self.get_serializer(data=request.data)
        serializer.is_valid(raise_exception=True)
        result = self.perform_create(serializer)
        # perform_create returns a Response only on the reuse-existing-booking path
        if isinstance(result, Response):
            return result
        instance = serializer.instance
        return Response(ExamBookingSerializer(instance, context={"request": request}).data, status=status.HTTP_201_CREATED)

    def perform_create(self, serializer):
        student_course = serializer.validated_data.get("student_course")
        if not student_course:
            raise ValidationError({"student_course": "This field is required."})

        exam = serializer.validated_data.get("exam")
        if not exam:
            raise ValidationError({"exam": "An exam list must be selected."})
        if exam.status == "closed":
            raise ValidationError({"exam": "Cannot add a student to a closed exam."})

        # Person 1 (HQ) assigns students who are pending_exam_booking
        if student_course.status != "pending_exam_booking":
            raise ValidationError(
                {"student_course": f"Student must be in 'Pending Exam' status (current: {student_course.status})."}
            )

        serializer.save(
            booked_by=self.request.user,
            student=student_course.student,
            admin1_comment=self.request.data.get("admin1_comment", ""),
        )
        student_course.status = "exam_list"
        student_course.save(update_fields=["status", "updated_at"])
        sync_student_status(student_course.student)

    @action(detail=True, methods=["post"])
    def approve(self, request, pk=None):
        booking = self.get_object()
        booking.status = "confirmed"
        booking.approved_by = request.user
        booking.approved_at = timezone.now()
        booking.admin2_comment = request.data.get("admin2_comment", booking.admin2_comment)
        booking.save()
        if booking.student_course:
            sc = booking.student_course
            if sc.status == "exam_list":
                sc.status = "exam_approved"
                sc.save(update_fields=["status", "updated_at"])
                sync_student_status(sc.student)
        return Response(ExamBookingSerializer(booking).data)

    @action(detail=True, methods=["post"])
    def remove_student(self, request, pk=None):
        """Remove a student from an exam (only while exam is still active, no result recorded)."""
        booking = self.get_object()

        if booking.exam and booking.exam.status == "closed":
            return Response(
                {"detail": "Cannot remove a student from a closed exam."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # Guard: never remove if a result already exists
        if hasattr(booking, "result"):
            return Response(
                {"detail": "Cannot remove a student whose result has been recorded."},
                status=status.HTTP_400_BAD_REQUEST,
            )

        sc = booking.student_course
        booking.delete()

        # Reset course back to pending_exam_booking so branch can rebook
        if sc and sc.status in ("exam_list", "exam_approved"):
            sc.status = "pending_exam_booking"
            sc.save(update_fields=["status", "updated_at"])
            sync_student_status(sc.student)

        return Response(status=status.HTTP_204_NO_CONTENT)


class ExamResultViewSet(viewsets.ModelViewSet):
    queryset = ExamResult.objects.select_related("exam_booking__student_course__student").all()
    serializer_class = ExamResultSerializer
    permission_classes = [IsAuthenticated]

    def perform_create(self, serializer):
        instance = serializer.save(recorded_by=self.request.user)
        self._apply_result(instance)

    def perform_update(self, serializer):
        instance = serializer.save()
        self._apply_result(instance)

    def _apply_result(self, result):
        sc = result.exam_booking.student_course
        booking = result.exam_booking
        if not sc:
            return
        if result.result in ("pass", "fail"):
            apply_exam_result(sc, result.result)
            if result.result == "fail":
                # Create a NEW pending booking for the retake — do NOT mutate the original.
                # The original booking stays attached to this exam for historical accuracy.
                ExamBooking.objects.create(
                    student_course=sc,
                    student=booking.student,
                    exam=None,
                    status="pending",
                    booked_by=booking.booked_by,
                )
