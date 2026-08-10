import io
from django.db.models import Sum
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
from django.http import HttpResponse
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Expense, ExpenseCategory
from .serializers import ExpenseSerializer, ExpenseCategorySerializer
from finance.models import PaymentTransaction
from branches.models import Branch


# ── Helpers ───────────────────────────────────────────────────────────────────

def _admin_required(user):
    if user.role != "super_admin":
        return Response({"detail": "Forbidden."}, status=403)
    return None


def _date_filter_expense(qs, params):
    if params.get("date_from"):
        qs = qs.filter(expense_date__gte=params["date_from"])
    if params.get("date_to"):
        qs = qs.filter(expense_date__lte=params["date_to"])
    return qs


def _date_filter_payment(qs, params):
    if params.get("date_from"):
        qs = qs.filter(created_at__date__gte=params["date_from"])
    if params.get("date_to"):
        qs = qs.filter(created_at__date__lte=params["date_to"])
    return qs


# ── Categories ────────────────────────────────────────────────────────────────

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def category_list_create(request):
    if request.method == "GET":
        qs = ExpenseCategory.objects.filter(is_active=True)
        return Response(ExpenseCategorySerializer(qs, many=True).data)
    serializer = ExpenseCategorySerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(["GET", "PUT", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def category_detail(request, pk):
    try:
        obj = ExpenseCategory.objects.get(pk=pk)
    except ExpenseCategory.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    if request.method == "GET":
        return Response(ExpenseCategorySerializer(obj).data)
    if request.method == "DELETE":
        obj.is_active = False
        obj.save()
        return Response(status=status.HTTP_204_NO_CONTENT)
    serializer = ExpenseCategorySerializer(obj, data=request.data, partial=request.method == "PATCH")
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


# ── Expenses CRUD ─────────────────────────────────────────────────────────────

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def expense_list_create(request):
    if request.method == "GET":
        qs = Expense.objects.select_related("branch", "category", "created_by")
        p = request.query_params
        if p.get("branch"):
            qs = qs.filter(branch_id=p["branch"])
        if p.get("expense_type"):
            qs = qs.filter(expense_type=p["expense_type"])
        if p.get("category"):
            qs = qs.filter(category_id=p["category"])
        if p.get("date_from"):
            qs = qs.filter(expense_date__gte=p["date_from"])
        if p.get("date_to"):
            qs = qs.filter(expense_date__lte=p["date_to"])
        return Response(ExpenseSerializer(qs, many=True).data)
    serializer = ExpenseSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    serializer.save(created_by=request.user)
    return Response(serializer.data, status=status.HTTP_201_CREATED)


@api_view(["GET", "PUT", "PATCH", "DELETE"])
@permission_classes([IsAuthenticated])
def expense_detail(request, pk):
    try:
        obj = Expense.objects.select_related("branch", "category", "created_by").get(pk=pk)
    except Expense.DoesNotExist:
        return Response(status=status.HTTP_404_NOT_FOUND)
    if request.method == "GET":
        return Response(ExpenseSerializer(obj).data)
    if request.method == "DELETE":
        obj.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)
    serializer = ExpenseSerializer(obj, data=request.data, partial=request.method == "PATCH")
    serializer.is_valid(raise_exception=True)
    serializer.save()
    return Response(serializer.data)


# ── Analytics ─────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def analytics_profitability_kpi(request):
    err = _admin_required(request.user)
    if err:
        return err

    params = request.query_params
    branch_id = params.get("branch")

    # Revenue — exclude transactions with no student (orphaned)
    rev_qs = _date_filter_payment(
        PaymentTransaction.objects.filter(status="completed", student__isnull=False), params
    )
    if branch_id:
        rev_qs = rev_qs.filter(student__branch_id=branch_id)
    total_revenue = rev_qs.aggregate(t=Sum("amount"))["t"] or 0

    exp_qs = _date_filter_expense(Expense.objects.all(), params)
    if branch_id:
        exp_qs = exp_qs.filter(branch_id=branch_id, expense_type="BRANCH")
    total_expenses = exp_qs.aggregate(t=Sum("amount"))["t"] or 0

    net_profit = float(total_revenue) - float(total_expenses)

    # Top branch by profit
    branch_rev = (
        _date_filter_payment(
            PaymentTransaction.objects.filter(status="completed", student__isnull=False), params
        )
        .values("student__branch_id", "student__branch__name")
        .annotate(rev=Sum("amount"))
    )
    rev_map = {r["student__branch_id"]: float(r["rev"] or 0) for r in branch_rev}
    exp_map = {
        e["branch_id"]: float(e["exp"] or 0)
        for e in _date_filter_expense(Expense.objects.filter(expense_type="BRANCH"), params)
        .values("branch_id").annotate(exp=Sum("amount"))
    }

    top_branch_name = top_profit = None
    if rev_map:
        best = max(set(rev_map) | set(exp_map), key=lambda bid: rev_map.get(bid, 0) - exp_map.get(bid, 0))
        for r in branch_rev:
            if r["student__branch_id"] == best:
                top_branch_name = r["student__branch__name"]
                top_profit = rev_map.get(best, 0) - exp_map.get(best, 0)
                break

    return Response({
        "total_revenue": float(total_revenue),
        "total_expenses": float(total_expenses),
        "net_profit": net_profit,
        "profit_margin": round((net_profit / float(total_revenue) * 100), 1) if total_revenue else 0,
        "top_branch": top_branch_name,
        "top_branch_profit": top_profit,
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def analytics_branch_profitability(request):
    err = _admin_required(request.user)
    if err:
        return err

    params = request.query_params
    branch_ids = [int(b) for b in params.getlist("branches") if b.isdigit()]

    rev_qs = _date_filter_payment(
        PaymentTransaction.objects.filter(status="completed", student__isnull=False), params
    )
    if branch_ids:
        rev_qs = rev_qs.filter(student__branch_id__in=branch_ids)
    rev_rows = (
        rev_qs
        .values("student__branch_id", "student__branch__name")
        .annotate(revenue=Sum("amount"))
        .order_by("student__branch__name")
    )

    exp_qs = _date_filter_expense(Expense.objects.filter(expense_type="BRANCH"), params)
    if branch_ids:
        exp_qs = exp_qs.filter(branch_id__in=branch_ids)
    exp_map = {
        e["branch_id"]: float(e["expense"] or 0)
        for e in exp_qs.values("branch_id").annotate(expense=Sum("amount"))
    }

    result = []
    for r in rev_rows:
        bid = r["student__branch_id"]
        revenue = float(r["revenue"] or 0)
        expense = exp_map.get(bid, 0)
        result.append({
            "branch_id": bid,
            "branch": r["student__branch__name"],
            "revenue": revenue,
            "expense": expense,
            "profit": revenue - expense,
            "margin": round(((revenue - expense) / revenue * 100), 1) if revenue else 0,
        })

    # Include branches with expenses but no revenue
    rev_ids = {r["branch_id"] for r in result}
    for bid, exp in exp_map.items():
        if bid not in rev_ids:
            try:
                branch = Branch.objects.get(pk=bid)
                result.append({"branch_id": bid, "branch": branch.name, "revenue": 0, "expense": exp, "profit": -exp, "margin": 0})
            except Branch.DoesNotExist:
                pass

    result.sort(key=lambda x: x["profit"], reverse=True)
    return Response(result)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def analytics_general_expenses(request):
    err = _admin_required(request.user)
    if err:
        return err

    params = request.query_params
    qs = _date_filter_expense(Expense.objects.filter(expense_type="GENERAL"), params)
    total = qs.aggregate(t=Sum("amount"))["t"] or 0
    categories = qs.values("category__name").annotate(value=Sum("amount")).order_by("-value")

    return Response({
        "total": float(total),
        "categories": [{"name": r["category__name"], "value": float(r["value"] or 0)} for r in categories],
    })


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def analytics_revenue_time_series(request):
    err = _admin_required(request.user)
    if err:
        return err

    params = request.query_params
    granularity = params.get("granularity", "monthly")
    TruncFn = {"daily": TruncDay, "weekly": TruncWeek, "monthly": TruncMonth}.get(granularity, TruncMonth)
    branch_id = params.get("branch")
    GENERAL = "__GENERAL__"

    rev_qs = _date_filter_payment(
        PaymentTransaction.objects.filter(status="completed", student__isnull=False), params
    )
    if branch_id and branch_id != GENERAL:
        rev_qs = rev_qs.filter(student__branch_id=branch_id)
    elif branch_id == GENERAL:
        rev_qs = rev_qs.none()

    rev_map = {
        (r["period"].date() if hasattr(r["period"], "date") else r["period"]): float(r["revenue"] or 0)
        for r in rev_qs
        .annotate(period=TruncFn("created_at"))
        .values("period")
        .annotate(revenue=Sum("amount"))
        .order_by("period")
    }

    exp_qs = _date_filter_expense(Expense.objects.all(), params)
    if branch_id == GENERAL:
        exp_qs = exp_qs.filter(expense_type="GENERAL")
    elif branch_id:
        exp_qs = exp_qs.filter(branch_id=branch_id, expense_type="BRANCH")
    elif params.get("expense_type"):
        exp_qs = exp_qs.filter(expense_type=params["expense_type"])

    exp_map = {
        (e["period"].date() if hasattr(e["period"], "date") else e["period"]): float(e["expense"] or 0)
        for e in exp_qs
        .annotate(period=TruncFn("expense_date"))
        .values("period")
        .annotate(expense=Sum("amount"))
        .order_by("period")
    }

    data = []
    for period in sorted(set(rev_map) | set(exp_map)):
        revenue = rev_map.get(period, 0)
        expense = exp_map.get(period, 0)
        data.append({
            "period": period.strftime("%Y-%m-%d"),
            "revenue": revenue,
            "expense": expense,
            "profit": revenue - expense,
        })

    return Response({"data": data, "granularity": granularity})


def _make_excel_response(wb, filename):
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    return response


def _style_header_row(ws, headers):
    fill = PatternFill("solid", fgColor="111827")
    font = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = width


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_profitability(request):
    err = _admin_required(request.user)
    if err:
        return err

    params = request.query_params

    # Branch rows
    rev_qs = _date_filter_payment(
        PaymentTransaction.objects.filter(status="completed", student__isnull=False), params
    )
    rev_rows = (
        rev_qs
        .values("student__branch_id", "student__branch__name")
        .annotate(revenue=Sum("amount"))
    )
    rev_map = {r["student__branch_id"]: {"name": r["student__branch__name"], "revenue": float(r["revenue"] or 0)} for r in rev_rows}

    exp_qs = _date_filter_expense(Expense.objects.filter(expense_type="BRANCH"), params)
    exp_map = {e["branch_id"]: float(e["expense"] or 0) for e in exp_qs.values("branch_id").annotate(expense=Sum("amount"))}

    rows = []
    for bid in set(rev_map) | set(exp_map):
        revenue = rev_map.get(bid, {}).get("revenue", 0)
        expense = exp_map.get(bid, 0)
        profit = revenue - expense
        rows.append({
            "branch": rev_map.get(bid, {}).get("name") or f"Branch #{bid}",
            "revenue": revenue,
            "expense": expense,
            "profit": profit,
            "margin": round((profit / revenue * 100), 1) if revenue else 0,
        })
    rows.sort(key=lambda x: x["profit"], reverse=True)

    # General operations row
    gen_total = float(
        _date_filter_expense(Expense.objects.filter(expense_type="GENERAL"), params)
        .aggregate(t=Sum("amount"))["t"] or 0
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profitability"
    headers = ["Branch", "Revenue (Ksh)", "Expenses (Ksh)", "Net Profit (Ksh)", "Margin %"]
    _style_header_row(ws, headers)

    for r in rows:
        ws.append([r["branch"], r["revenue"], r["expense"], r["profit"], r["margin"]])
    if gen_total > 0:
        ws.append(["General Operations", 0, gen_total, -gen_total, "—"])

    _autofit(ws)
    return _make_excel_response(wb, "profitability.xlsx")


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_expenses(request):
    qs = Expense.objects.select_related("branch", "category", "created_by")
    p = request.query_params
    if p.get("branch"):
        qs = qs.filter(branch_id=p["branch"])
    if p.get("expense_type"):
        qs = qs.filter(expense_type=p["expense_type"])
    if p.get("date_from"):
        qs = qs.filter(expense_date__gte=p["date_from"])
    if p.get("date_to"):
        qs = qs.filter(expense_date__lte=p["date_to"])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Expenses"
    headers = ["Date", "Type", "Category", "Description", "Branch", "Amount (Ksh)", "Created By"]
    _style_header_row(ws, headers)

    for e in qs:
        ws.append([
            str(e.expense_date),
            e.expense_type,
            e.category.name if e.category else "",
            e.description,
            e.branch.name if e.branch else "—",
            float(e.amount),
            e.created_by.get_full_name() if e.created_by else "",
        ])

    _autofit(ws)
    return _make_excel_response(wb, "expenses.xlsx")
