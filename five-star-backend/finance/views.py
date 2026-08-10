from decimal import Decimal, InvalidOperation
from typing import Tuple
import logging
import io
from rest_framework import viewsets, status
from rest_framework.decorators import permission_classes, api_view
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.response import Response
from django.db.models import Prefetch, Q, Sum, Count
from django.db import IntegrityError
from django.conf import settings
from django.http import HttpResponse
from django.utils.dateparse import parse_date, parse_datetime
from django.utils import timezone

from academics.models import StudentCourse
from .models import PaymentTransaction, PaymentConfirmation, MpesaSTKRequest, PaymentValidation, BankNotification
from .serializers import PaymentTransactionSerializer, StudentCourseSearchSerializer


logger = logging.getLogger('payments')


def _contact_footer(branch):
    lines = []
    if branch:
        branch_line = f"{branch.name} Branch"
        if branch.phone_number:
            branch_line += f": {branch.phone_number}"
        lines.append(branch_line)
    lines.append("HQ: +254 727 555 558")
    return "\n".join(lines)


def _extract_mpesa_reference(narration: str) -> str:
    if not narration:
        return ""
    part = narration.split("~")[0].strip()
    return part if part else ""


def _compute_balance_snapshot(student_course, new_payment_amount):
    """
    Returns (previous_balance, new_balance) at the moment a payment is being created.
    Must be called BEFORE the new PaymentTransaction is saved.
    """
    if not student_course:
        return None, None
    from django.db.models import Sum
    agreed = student_course.amount_agreed or 0
    paid_so_far = student_course.payments.filter(status="completed").aggregate(
        total=Sum("amount")
    )["total"] or 0
    previous_balance = agreed - paid_so_far
    new_balance = previous_balance - new_payment_amount
    return previous_balance, new_balance


def _classify_payment_type(student_course):
    if not student_course:
        return "UNALLOCATED"
    sc_status = student_course.status
    if sc_status == "retake_booked":
        return "RETAKE"
    if sc_status == "dormant":
        return "PDL_REACTIVATION"
    has_prior = PaymentTransaction.objects.filter(
        student_course=student_course,
        status="completed",
    ).exists()
    return "TOP_UP" if has_prior else "REGISTRATION"


class StandardPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 100

    def get_paginated_response(self, data):
        return Response({
            "count": self.page.paginator.count,
            "total_pages": self.page.paginator.num_pages,
            "results": data,
        })


class PaymentTransactionViewSet(viewsets.ModelViewSet):
    serializer_class = PaymentTransactionSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardPagination

    def _base_queryset(self):
        return PaymentTransaction.objects.select_related(
            "branch",
            "student__branch",
            "student_course__student",
            "student_course__course",
        ).prefetch_related(
            Prefetch("student_course__payments",
                     queryset=PaymentTransaction.objects.only("student_course_id", "amount"))
        ).all()

    def _apply_filters(self, qs, user, params):
        if user.role == "branch_user":
            qs = qs.filter(Q(student__branch=user.branch) | Q(branch=user.branch))
        elif user.role == "super_admin":
            branch_id = params.get("branch_id")
            if branch_id:
                qs = qs.filter(Q(student__branch_id=branch_id) | Q(branch_id=branch_id))

        if params.get("student_course_id"):
            qs = qs.filter(student_course_id=params["student_course_id"])
        if params.get("student_id"):
            qs = qs.filter(student_course__student_id=params["student_id"])
        if params.get("status"):
            qs = qs.filter(status=params["status"])
        if params.get("payment_type"):
            qs = qs.filter(payment_type=params["payment_type"])
        if params.get("channel"):
            qs = qs.filter(channel=params["channel"])

        date_from = params.get("date_from")
        date_to = params.get("date_to")
        if date_from:
            qs = qs.filter(transaction_date__date__gte=parse_date(date_from))
        if date_to:
            qs = qs.filter(transaction_date__date__lte=parse_date(date_to))

        search = params.get("search")
        if search:
            qs = qs.filter(
                Q(reference_code__icontains=search)
                | Q(mpesa_reference__icontains=search)
                | Q(student__full_name__icontains=search)
                | Q(student__admission_number__icontains=search)
            )
        return qs

    def get_queryset(self):
        qs = self._base_queryset()
        qs = self._apply_filters(qs, self.request.user, self.request.query_params)
        from django.db.models import F
        return qs.order_by(F("transaction_date").desc(nulls_last=True))

    def get_serializer_class(self):
        if self.action == "retrieve":
            from .serializers import PaymentTransactionDetailSerializer
            return PaymentTransactionDetailSerializer
        return PaymentTransactionSerializer

    def perform_create(self, serializer):
        student_course = serializer.validated_data.get("student_course")
        student = student_course.student if student_course else None
        payment_type = _classify_payment_type(student_course)
        amount = serializer.validated_data.get("amount", 0)
        prev_bal, new_bal = _compute_balance_snapshot(student_course, amount)
        serializer.save(
            student=student,
            payment_type=payment_type,
            receipt_previous_balance=prev_bal,
            receipt_new_balance=new_bal,
        )

    def partial_update(self, request, *args, **kwargs):
        return super().partial_update(request, *args, **kwargs)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def search_student_course(request):
    if request.user.role not in ("super_admin",):
        return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)
    ref = request.query_params.get("payment_reference", "").strip()
    if not ref:
        return Response([], status=status.HTTP_200_OK)
    results = StudentCourse.objects.select_related("student", "course").filter(
        Q(payment_reference__icontains=ref) | Q(student__full_name__icontains=ref)
    )[:10]
    return Response(StudentCourseSearchSerializer(results, many=True).data)


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def allocate_payment(request, pk):
    if request.user.role not in ("super_admin",):
        return Response({"detail": "Forbidden"}, status=status.HTTP_403_FORBIDDEN)
    try:
        tx = PaymentTransaction.objects.get(pk=pk, status="orphaned")
    except PaymentTransaction.DoesNotExist:
        return Response({"detail": "Not found or not orphaned"}, status=status.HTTP_404_NOT_FOUND)
    sc_id = request.data.get("student_course_id")
    if not sc_id:
        return Response({"detail": "student_course_id required"}, status=status.HTTP_400_BAD_REQUEST)
    try:
        sc = StudentCourse.objects.select_related("student").get(pk=sc_id)
    except StudentCourse.DoesNotExist:
        return Response({"detail": "StudentCourse not found"}, status=status.HTTP_404_NOT_FOUND)
    tx.student_course = sc
    tx.student = sc.student
    tx.branch = None
    tx.status = "completed"
    tx.payment_type = _classify_payment_type(sc)
    prev_bal, new_bal = _compute_balance_snapshot(sc, tx.amount)
    tx.receipt_previous_balance = prev_bal
    tx.receipt_new_balance = new_bal
    tx.save(update_fields=["student_course", "student", "branch", "status", "payment_type",
                           "receipt_previous_balance", "receipt_new_balance", "updated_at"])
    # lifecycle transition handled by finance.signals.payment_transaction_post_save
    return Response({"detail": "Allocated successfully"})


def coop_auth(connectionID, connectionPassword):
    if settings.CONNECTION_ID != connectionID:
        return False

    if settings.CONNECTION_PASSWORD != connectionPassword:
        return False

    return True


@api_view(["POST"])
@permission_classes([AllowAny])
def payment_validation(request):
    logger.info(f"Received Payment validation request {request.data}")

    data = request.data

    pv = PaymentValidation.record(data)
    logger.info(f"Payment validation object created {pv}")

    header = data.get("header", {})
    req = data.get("request", {})

    # auth
    connectionID = header.get("connectionID")
    connectionPassword = header.get("connectionPassword")

    auth_success = coop_auth(connectionID, connectionPassword)
    if not auth_success:
        logger.info(f"Payment validation auth failed id: {connectionID} password: {connectionPassword}")
        return Response({}, status=status.HTTP_400_BAD_REQUEST)

    message_id = header.get("messageID", "")
    service_name = header.get("serviceName", "")

    transaction_reference = req.get("TransactionReferenceCode", "")
    transaction_date = req.get("TransactionDate", "")
    institution_code = req.get("InstitutionCode", "")

    account_name = "Student"

    response_data = {
        "header": {
            "messageID": message_id,
            "statusCode": "200",
            "statusDescription": "Successfully validated customer",
        },
        "response": {
            "TransactionReferenceCode": transaction_reference,
            "TransactionDate": transaction_date,
            "TotalAmount": 0,
            "Currency": "KES",
            "AdditionalInfo": account_name,
            "AccountNumber": transaction_reference,
            "AccountName": account_name,
            "InstitutionCode": institution_code,
            "InstitutionName": service_name,
        },
    }

    return Response(response_data, status=status.HTTP_200_OK)


@api_view(["POST"])
@permission_classes([AllowAny])
def payment_confirmation(request):
    logger.info(f"Received Payment confirmation request {request.data}")

    data = request.data
    pc = PaymentConfirmation.record(data)
    logger.info(f"Payment confirmation object created {pc}")

    header = data.get("header", {})
    req = data.get("request", {})

    # auth
    connectionID = header.get("connectionID")
    connectionPassword = header.get("connectionPassword")

    auth_success = coop_auth(connectionID, connectionPassword)
    if not auth_success:
        logger.info(f"Payment confirmation auth failed id: {connectionID} password: {connectionPassword}")
        return Response({}, status=status.HTTP_400_BAD_REQUEST)

    message_id = header.get("messageID", "")

    transaction_reference = req.get("TransactionReferenceCode", "")
    transaction_date = req.get("TransactionDate", "")
    payment_amount = req.get("PaymentAmount", req.get("TotalAmount", "0"))
    total_amount = req.get("TotalAmount", payment_amount)
    account_number = req.get("AccountNumber", "")
    institution_code = req.get("InstitutionCode", "")
    institution_name = req.get("InstitutionName", "")
    currency = req.get("Currency", "KES")
    doc_reference = req.get("DocumentReferenceNumber")
    amount = Decimal(str(req.get("PaymentAmount", req.get("TotalAmount", "0"))))

    student_course, branch, payment_status = _resolve_payment_target(
        (doc_reference or "").strip().upper()
    )
    student = student_course.student if student_course else None

    payment_tx = PaymentTransaction.objects.filter(
        reference_code=transaction_reference
    ).first()

    if not payment_tx:
        prev_bal, new_bal = _compute_balance_snapshot(student_course, amount)
        payment_tx = PaymentTransaction.objects.create(
            student_course=student_course,
            student=student,
            branch=branch,
            amount=amount,
            reference_code=transaction_reference,
            payment_method="bank_b2b",
            status=payment_status,
            payment_type=_classify_payment_type(student_course) if payment_status == "completed" else "UNALLOCATED",
            description=f"Payment received for account {doc_reference}",
            data=data,
            receipt_previous_balance=prev_bal,
            receipt_new_balance=new_bal,
        )
        pc.payment_transaction = payment_tx
        pc.save()
        logger.info(f"Payment validation transaction created {payment_tx}")

    response_data = {
        "header": {
            "messageID": message_id,
            "statusCode": "200",
            "statusDescription": "Payment successfully received",
        },
        "response": {
            "TransactionReferenceCode": transaction_reference,
            "TransactionDate": transaction_date,
            "TransactionAmount": payment_amount,
            "AccountNumber": account_number,
            "AccountName": "",
            "InstitutionCode": institution_code,
            "InstitutionName": institution_name,
            "Currency": currency,
            "AdditionalInfo": account_number,
            "TotalAmount": total_amount,
        },
    }

    return Response(response_data, status=status.HTTP_200_OK)

def extract_narration_identifiers(narration: str) -> dict:
    ref_code = ""
    mpesa_reference = ""
    channel = "UNKNOWN"

    try:
        if "MPESAC2B" in narration:
            channel = "MPESA"
            parts = narration.split("~")
            mpesa_reference = parts[0]
            ref_code = parts[1].split("#")[1] if len(parts) > 1 else ""

        elif "payment_reference" in narration:
            channel = "STK PUSH"
            parts = narration.split("~")
            mpesa_reference = parts[0]
            ref_code = parts[1] if len(parts) > 1 else ""

        elif "PESALINK" in narration:
            channel = "PESALINK"
            ref_code = narration.split("~")[-1]

        elif "POSAG" in narration:
            channel = "AGENT DEPOSIT"
            # Format: POSAG037661 ~616213022705~SUK001~POS39673_01192245021100
            # The payment reference is the 3rd tilde-delimited segment (index 2)
            parts = [p.strip() for p in narration.split("~")]
            ref_code = parts[2] if len(parts) > 2 else ""

    except Exception:
        pass

    return {
        "ref_code": ref_code.strip().upper(),
        "channel": channel,
        "mpesa_reference": mpesa_reference.strip(),
    }


def _resolve_payment_target(ref_code: str):
    """
    Returns (student_course, branch, status).
    Lookup order:
      1. Exact StudentCourse match  → completed
      2. Branch code match          → orphaned with branch
      3. No match                   → orphaned, no branch
    Never auto-allocates on a branch-only match.
    """
    from branches.models import Branch

    if ref_code:
        sc = StudentCourse.objects.filter(payment_reference=ref_code).first()
        if sc:
            return sc, None, "completed"

        branch = Branch.objects.filter(branch_code=ref_code).first()
        if branch:
            return None, branch, "orphaned"

    return None, None, "orphaned"

@api_view(["POST"])
@permission_classes([AllowAny])
def bank_payment_notification(request):
    logger.info(f"Received Bank Payment notification request {request.data}")

    data = request.data
    bn = BankNotification.record(data)
    logger.info(f"Received Bank Payment notification request {bn}")

    transaction_id = data.get("TransactionId", "")
    reference_code = data.get("PaymentRef", "")
    amount_raw = data.get("Amount", "0")
    currency = data.get("Currency", "KES")
    event_type = data.get("EventType", "")
    transaction_date = data.get("TransactionDate", "")
    posting_date = data.get("PostingDate", "")
    narration = data.get("Narration", "")
    acct_no = data.get("AcctNo", "")

    try:
        amount = Decimal(str(amount_raw))
    except (InvalidOperation, TypeError, ValueError):
        logger.info(f"Payment Bank notification failed due to invalid amount {request.data}")
        return Response(
            {
                "MessageCode": "400",
                "Message": "Invalid Amount",
            },
            status=status.HTTP_400_BAD_REQUEST,
        )

    if event_type.upper() != "CREDIT":
        logger.info(f"Payment Bank notification for non credit event {event_type}")
        return Response(
            {
                "MessageCode": "200",
                "Message": "Successfully received data",
            },
            status=status.HTTP_200_OK,
        )

    narration_data = extract_narration_identifiers(narration)
    ref_code = narration_data.get("ref_code")
    channel = narration_data.get("channel")
    mpesa_ref = narration_data.get("mpesa_reference")

    student_course, branch, payment_status = _resolve_payment_target(ref_code)
    student = student_course.student if student_course else None

    payment_tx = PaymentTransaction.objects.filter(
        reference_code=transaction_id
    ).first()

    if not payment_tx:
        payment_timestamp = (
            parse_datetime(transaction_date) if transaction_date else None
        ) or (
            parse_datetime(posting_date) if posting_date else None
        ) or timezone.now()

        prev_bal, new_bal = _compute_balance_snapshot(student_course, amount)
        payment_tx = PaymentTransaction.objects.create(
            student_course=student_course,
            student=student,
            branch=branch,
            amount=amount,
            reference_code=transaction_id,
            payment_method="bank_ipn",
            status=payment_status,
            channel=channel,
            payment_type=_classify_payment_type(student_course) if payment_status == "completed" else "UNALLOCATED",
            mpesa_reference=mpesa_ref or None,
            transaction_date=payment_timestamp,
            description=(
                f"Bank payment received. "
                f"Account: {acct_no}, "
                f"PaymentRef: {reference_code}, "
                f"Currency: {currency}, "
                f"TransactionDate: {transaction_date or posting_date}, "
                f"Narration: {narration}"
            ),
            data=data,
            receipt_previous_balance=prev_bal,
            receipt_new_balance=new_bal,
        )
        bn.payment_transaction = payment_tx
        bn.save()
        logger.info(f"Payment transaction for Bank Notification created {payment_tx}")

    return Response(
        {
            "MessageCode": "200",
            "Message": "Successfully received data",
        },
        status=status.HTTP_200_OK,
    )

from decimal import Decimal


def _validate_coop_phone(phone_raw: str) -> Tuple[bool, str]:
    phone = phone_raw.strip().replace(" ", "")
    # Accept +2547... or 2547... and convert back to 07 for Co-op
    if phone.startswith("+254"):
        phone = "0" + phone[4:]
    elif phone.startswith("254") and len(phone) == 12:
        phone = "0" + phone[3:]
    if len(phone) != 10 or not (phone.startswith("07") or phone.startswith("01")):
        return False, "Invalid phone number. Use format 07XXXXXXXX."
    return True, phone


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def coop_stk_push(request):
    """Initiate a Co-op STK Push for a StudentCourse."""
    from services.coop_service import initiate_stk_push, make_message_reference
    from django.utils import timezone as tz
    from datetime import timedelta

    student_course_id = request.data.get("student_course_id")
    phone_raw         = request.data.get("phone", "").strip()
    amount_raw        = request.data.get("amount")

    if not all([student_course_id, phone_raw, amount_raw]):
        return Response(
            {"detail": "student_course_id, phone, and amount are required."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    try:
        amount = int(amount_raw)
        if amount <= 0:
            raise ValueError
    except (ValueError, TypeError):
        return Response({"detail": "Invalid amount."}, status=status.HTTP_400_BAD_REQUEST)

    # Co-op expects 07XXXXXXXX — validate and normalise accordingly
    valid, phone = _validate_coop_phone(phone_raw)
    if not valid:
        return Response({"detail": phone}, status=status.HTTP_400_BAD_REQUEST)

    try:
        sc = StudentCourse.objects.select_related("student").get(pk=student_course_id)
    except StudentCourse.DoesNotExist:
        return Response({"detail": "Student course not found."}, status=status.HTTP_404_NOT_FOUND)

    payment_reference = sc.payment_reference
    if not payment_reference:
        return Response(
            {"detail": "This course has no payment reference."},
            status=status.HTTP_400_BAD_REQUEST,
        )

    # Expire stale pending requests older than 2 minutes (callback/enquiry never resolved them)
    MpesaSTKRequest.objects.filter(
        account_reference=payment_reference,
        status="pending",
        created_at__lt=tz.now() - timedelta(minutes=2),
    ).update(status="timeout")

    if MpesaSTKRequest.objects.filter(account_reference=payment_reference, status="pending").exists():
        return Response(
            {"detail": "A payment request is already pending for this course."},
            status=status.HTTP_409_CONFLICT,
        )

    message_reference = make_message_reference(payment_reference)

    try:
        stk_data = initiate_stk_push(
            message_reference=message_reference,
            payment_reference=payment_reference,
            phone=phone,        # 07XXXXXXXX as Co-op expects
            amount=amount,
        )
    except Exception as exc:
        logger.error(f"Co-op STK Push failed: {exc}")
        return Response(
            {"detail": "Failed to initiate Co-op payment. Please try again."},
            status=status.HTTP_502_BAD_GATEWAY,
        )

    MpesaSTKRequest.objects.create(
        checkout_request_id=message_reference,
        merchant_request_id="",
        phone_number=phone,
        amount=amount,
        account_reference=payment_reference,
    )

    return Response(
        {
            "checkout_request_id": message_reference,
            "message": "STK Push sent. Please check your phone and enter your Co-op PIN.",
        },
        status=status.HTTP_200_OK,
    )

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def coop_stk_status(request, checkout_request_id):

    try:
        stk = MpesaSTKRequest.objects.get(checkout_request_id=checkout_request_id)
    except MpesaSTKRequest.DoesNotExist:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    # ── Already resolved — return immediately, zero API calls ─────────────────
    if stk.status != "pending":
        return Response({"status": stk.status, "result_desc": stk.result_desc})

    is_final = request.query_params.get("final") == "1"
    poll_num = int(request.query_params.get("poll", "1"))

    if not is_final and poll_num <= 3:
        return Response({"status": "pending", "result_desc": "Waiting for customer response…"})

    # ── Phase B / final: call Co-op enquiry ───────────────────────────────────
    from services.coop_service import get_stk_status, interpret_enquiry_response

    try:
        raw = get_stk_status(checkout_request_id)   # checkout_request_id IS the MessageReference
    except Exception as exc:
        logger.error(f"Co-op STK enquiry error for {checkout_request_id} (poll={poll_num}): {exc}")
        if is_final:
            stk.status      = "timeout"
            stk.result_desc = "Co-op enquiry unreachable; outcome unknown."
            stk.save(update_fields=["status", "result_desc", "updated_at"])
            return Response({"status": "timeout", "result_desc": stk.result_desc})
        # Transient network error — keep retrying.
        return Response({"status": "pending", "result_desc": "Checking payment status…"})

    logger.info(f"Co-op STK enquiry response for {checkout_request_id} (poll={poll_num}): {raw}")
    interpreted    = interpret_enquiry_response(raw)
    enquiry_status = interpreted["status"]
    result_desc    = interpreted["result_desc"]
    txn_ref        = interpreted["transaction_reference"]

    # ── Still processing ──────────────────────────────────────────────────────
    if enquiry_status in ("pending", "unknown"):
        if is_final:
            stk.status      = "timeout"
            stk.result_desc = result_desc or "No response received within the polling window."
            stk.save(update_fields=["status", "result_desc", "updated_at"])
            return Response({"status": "timeout", "result_desc": stk.result_desc})
        return Response({"status": "pending", "result_desc": result_desc or "Processing…"})

    # ── Success ───────────────────────────────────────────────────────────────
    if enquiry_status == "success":
        stk.status        = "success"
        stk.result_desc   = result_desc
        stk.mpesa_receipt = txn_ref or stk.mpesa_receipt
        stk.save(update_fields=["status", "result_desc", "mpesa_receipt", "updated_at"])
        return Response({"status": "success", "result_desc": result_desc})

    # ── Failed / cancelled ────────────────────────────────────────────────────
    stk.status      = "failed"
    stk.result_desc = result_desc
    stk.save(update_fields=["status", "result_desc", "updated_at"])
    return Response({"status": "failed", "result_desc": result_desc})


@api_view(["POST", "GET"])
@permission_classes([AllowAny])
def coop_stk_callback(request):
    """
    Co-op STK Push callback — updates MpesaSTKRequest state only.
    Payment creation is handled exclusively by the IPN service.
    """
    import json
    data    = request.data if request.data else {}
    headers = dict(request.headers)
    logger.info(f"Co-op STK callback received: method={request.method} body={data}")

    try:
        with open("logs/coop_raw_capture.log", "a") as f:
            f.write(json.dumps({"method": request.method, "headers": headers, "body": data}) + "\n")
    except Exception as e:
        logger.error(f"Co-op STK callback log write failed: {e}")

    message_reference = data.get("MessageReference", "")
    if not message_reference:
        return Response({"status": "accepted"}, status=200)

    try:
        stk = MpesaSTKRequest.objects.get(checkout_request_id=message_reference)
    except MpesaSTKRequest.DoesNotExist:
        logger.warning(f"Co-op STK callback for unknown MessageReference: {message_reference}")
        return Response({"status": "accepted"}, status=200)

    if stk.status != "pending":
        return Response({"status": "accepted"}, status=200)

    result_code = str(data.get("ResponseCode", data.get("StatusCode", "")))
    result_desc = data.get("ResponseDescription", data.get("StatusDescription", ""))
    txn_ref     = data.get("TransactionReference", data.get("TransactionID", ""))

    new_status        = "success" if result_code == "0" else "failed"
    stk.status        = new_status
    stk.result_code   = result_code
    stk.result_desc   = result_desc
    stk.mpesa_receipt = txn_ref
    stk.save(update_fields=["status", "result_code", "result_desc", "mpesa_receipt", "updated_at"])
    logger.info(f"Co-op STK callback updated {message_reference} → {new_status}")

    return Response({"status": "accepted"}, status=200)



@api_view(["GET"])
@permission_classes([IsAuthenticated])
def payments_summary(request):
    """Returns aggregated payment stats for the current filters (no pagination)."""
    viewset = PaymentTransactionViewSet()
    viewset.request = request
    qs = viewset._base_queryset()
    qs = viewset._apply_filters(qs, request.user, request.query_params)

    rows = qs.values("payment_type", "status").annotate(
        count=Count("id"), revenue=Sum("amount")
    )

    result = {"total_count": 0, "total_revenue": 0, "by_type": {}}
    for row in rows:
        pt = row["payment_type"]
        result["total_count"] += row["count"]
        result["total_revenue"] += float(row["revenue"] or 0)
        if pt not in result["by_type"]:
            result["by_type"][pt] = {"count": 0, "revenue": 0, "completed": 0, "orphaned": 0}
        result["by_type"][pt]["count"] += row["count"]
        result["by_type"][pt]["revenue"] += float(row["revenue"] or 0)
        if row["status"] in ("completed", "orphaned"):
            result["by_type"][pt][row["status"]] += row["count"]

    return Response(result)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_payments(request):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    viewset = PaymentTransactionViewSet()
    viewset.request = request
    qs = viewset._base_queryset()
    qs = viewset._apply_filters(qs, request.user, request.query_params)
    from django.db.models import F
    qs = qs.order_by(F("transaction_date").desc(nulls_last=True))

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Payments"

    headers = ["Reference", "Student Name", "Admission", "Course", "Branch",
               "Amount", "Channel", "Payment Type", "Status", "Payment Method", "Date", "Balance"]

    header_fill = PatternFill("solid", fgColor="111827")
    header_font = Font(bold=True, color="FFFFFF", size=10)

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row_idx, p in enumerate(qs, 2):
        sc = p.student_course
        balance = ""
        if sc:
            agreed = float(sc.amount_agreed or 0)
            total_paid = float(sum(x.amount for x in sc.payments.filter(status="completed")))
            balance = round(agreed - total_paid, 2)
        ws.append([
            p.reference_code,
            p.student.full_name if p.student else "",
            p.student.admission_number if p.student else "",
            sc.course.class_name if sc and sc.course else "",
            p.student.branch.name if p.student and p.student.branch else "",
            float(p.amount),
            p.channel or "",
            p.get_payment_type_display(),
            p.status.capitalize(),
            p.payment_method.capitalize(),
            (p.transaction_date or p.created_at).strftime("%Y-%m-%d %H:%M") if (p.transaction_date or p.created_at) else "",
            balance,
        ])

    for col in ws.columns:
        ws.column_dimensions[col[0].column_letter].width = max(
            len(str(col[0].value or "")), max((len(str(c.value or "")) for c in col[1:]), default=0)
        ) + 4

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    response = HttpResponse(
        buffer.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="payments_export.xlsx"'
    return response


@api_view(["POST"])
@permission_classes([IsAuthenticated])
def record_print(request, pk):
    """Record a receipt print. Branch users limited to 1 action (2 physical copies).
    Super admins can always print."""
    try:
        tx = PaymentTransaction.objects.get(pk=pk)
    except PaymentTransaction.DoesNotExist:
        return Response({"detail": "Not found."}, status=status.HTTP_404_NOT_FOUND)

    # Enforce limit for non-super-admins
    if request.user.role != "super_admin":
        if tx.receipt_print_count >= 1:
            return Response(
                {"detail": "Receipt already printed. Maximum prints reached."},
                status=status.HTTP_403_FORBIDDEN,
            )
        # Branch user must own this payment
        if request.user.role == "branch_user" and tx.student and tx.student.branch != request.user.branch:
            return Response({"detail": "Forbidden."}, status=status.HTTP_403_FORBIDDEN)

    from django.utils import timezone
    tx.receipt_print_count += 1
    if not tx.first_printed_at:
        tx.first_printed_at = timezone.now()
        tx.printed_by = request.user
    tx.save(update_fields=["receipt_print_count", "first_printed_at", "printed_by", "updated_at"])

    return Response(PaymentTransactionSerializer(tx).data)


