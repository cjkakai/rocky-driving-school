"""
reports/views.py

Report is now a lightweight daily sign-off: branch, created_by, report_date,
inquiries (the one genuinely manual input), and an optional note. Everything
else — registrations, payments, enrollments, exam bookings — is computed
live from the source models on every read, never frozen at submission time.

Vehicle/trip data has moved to Lesson (academics app); Report no longer
tracks it. TripEntry is legacy-only and is not written to or read from here.
"""

from django.db.models import Sum, Count, Q
from django.db.models.functions import TruncDay, TruncWeek, TruncMonth
from django.utils import timezone
from django.http import HttpResponse
from datetime import datetime, time, timedelta
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response

from students.models import Student
from finance.models import PaymentTransaction
from bookings.models import ExamBooking
from academics.models import StudentCourse, Course
from branches.models import Branch

from .models import Report
from .serializers import ReportSerializer


# ── Date helpers ─────────────────────────────────────────────────────────────

def _parse_date(s):
    return datetime.strptime(s, "%Y-%m-%d").date()

def _day_range(d):
    start = timezone.make_aware(datetime.combine(d, time.min))
    end   = timezone.make_aware(datetime.combine(d, time.max))
    return start, end

def _auto_granularity(date_from, date_to):
    delta = (date_to - date_from).days
    if delta <= 14: return "daily"
    if delta <= 90: return "weekly"
    return "monthly"


# ── Filter helpers ────────────────────────────────────────────────────────────

def _dt_from(params):
    s = params.get("date_from")
    return timezone.make_aware(datetime.combine(_parse_date(s), time.min)) if s else None

def _dt_to(params):
    s = params.get("date_to")
    return timezone.make_aware(datetime.combine(_parse_date(s), time.max)) if s else None

def _apply_dt_filters(qs, params, field="created_at"):
    dt_from = _dt_from(params)
    dt_to   = _dt_to(params)
    if dt_from: qs = qs.filter(**{f"{field}__gte": dt_from})
    if dt_to:   qs = qs.filter(**{f"{field}__lte": dt_to})
    return qs

def _apply_date_filters_report(qs, params):
    if params.get("date_from"): qs = qs.filter(report_date__gte=params["date_from"])
    if params.get("date_to"):   qs = qs.filter(report_date__lte=params["date_to"])
    return qs

def _apply_date_filters_sc(qs, params):
    if params.get("date_from"): qs = qs.filter(registration_date__gte=params["date_from"])
    if params.get("date_to"):   qs = qs.filter(registration_date__lte=params["date_to"])
    return qs

def _bf(qs, field, branch_id):
    return qs.filter(**{field: branch_id}) if branch_id else qs


# ── Auto metrics (live, for preview + daily_summary) ─────────────────────────

def _compute_auto_metrics(branch_id, report_date):
    start, end = _day_range(report_date)
    student_regs = Student.objects.filter(branch_id=branch_id, created_at__range=[start, end]).count()
    sc_regs      = StudentCourse.objects.filter(student__branch_id=branch_id, registration_date__range=[start, end]).count()
    payments     = PaymentTransaction.objects.filter(
        student__branch_id=branch_id, status="completed", created_at__range=[start, end]
    ).aggregate(count=Count("id"), total=Sum("amount"))
    exam_count   = ExamBooking.objects.filter(student__branch_id=branch_id, created_at__range=[start, end]).count()
    course_rows  = StudentCourse.objects.filter(
        student__branch_id=branch_id, registration_date__range=[start, end]
    ).values("course__class_name").annotate(cnt=Count("id"))
    return {
        "student_registrations":        student_regs,
        "student_course_registrations": sc_regs,
        "payment_count":                payments["count"] or 0,
        "payment_total":                float(payments["total"] or 0),
        "exam_bookings_count":          exam_count,
        "course_breakdown":             {r["course__class_name"]: r["cnt"] for r in course_rows},
    }


# ── CRUD ──────────────────────────────────────────────────────────────────────

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def reports_list_create(request):
    user = request.user

    if request.method == "GET":
        qs = Report.objects.select_related("branch", "created_by")
        if user.role == "branch_user":
            qs = qs.filter(branch=user.branch)
        else:
            if request.query_params.get("branch"):
                qs = qs.filter(branch_id=int(request.query_params["branch"]))
            if request.query_params.get("date_from"):
                qs = qs.filter(report_date__gte=request.query_params["date_from"])
            if request.query_params.get("date_to"):
                qs = qs.filter(report_date__lte=request.query_params["date_to"])
        return Response(ReportSerializer(qs, many=True).data)

    # ── POST ──────────────────────────────────────────────────────────────────
    if user.role == "branch_user" and not user.branch:
        return Response({"detail": "No branch assigned."}, status=400)

    branch = user.branch if user.role == "branch_user" else None
    if user.role == "super_admin":
        branch_id_raw = request.data.get("branch")
        if not branch_id_raw:
            return Response({"detail": "branch is required."}, status=400)
        try:
            branch = Branch.objects.get(id=int(branch_id_raw))
        except Branch.DoesNotExist:
            return Response({"detail": "Branch not found."}, status=404)

    try:
        report_date = _parse_date(request.data["report_date"])
    except (KeyError, ValueError):
        return Response({"detail": "report_date (YYYY-MM-DD) is required."}, status=400)

    today     = timezone.now().date()
    yesterday = today - timedelta(days=1)
    if user.role == "branch_user" and report_date not in (today, yesterday):
        return Response({"detail": "You can only submit a report for today or yesterday."}, status=400)
    if report_date > today:
        return Response({"detail": "Cannot submit a report for a future date."}, status=400)
    if Report.objects.filter(branch=branch, report_date=report_date).exists():
        return Response({"detail": f"A report for {report_date} already exists."}, status=400)

    report = Report.objects.create(
        branch=branch,
        created_by=user,
        report_date=report_date,
        inquiries=int(request.data.get("inquiries", 0)),
        notes=request.data.get("notes", "").strip(),
    )
    return Response(ReportSerializer(report).data, status=201)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_detail(request, pk):
    try:
        report = Report.objects.select_related("branch", "created_by").get(pk=pk)
    except Report.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if request.user.role == "branch_user" and report.branch != request.user.branch:
        return Response({"detail": "Forbidden."}, status=403)
    return Response(ReportSerializer(report).data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_preview(request):
    user = request.user
    try:
        report_date = _parse_date(request.query_params["report_date"])
    except (KeyError, ValueError):
        return Response({"detail": "report_date (YYYY-MM-DD) required."}, status=400)
    branch_id = user.branch_id if user.role == "branch_user" else request.query_params.get("branch")
    if not branch_id:
        return Response({"detail": "branch required."}, status=400)
    already_exists = Report.objects.filter(branch_id=int(branch_id), report_date=report_date).exists()
    data = _compute_auto_metrics(int(branch_id), report_date)
    data["already_submitted"] = already_exists
    return Response(data)


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def report_drilldown(request, pk):
    try:
        report = Report.objects.select_related("branch").get(pk=pk)
    except Report.DoesNotExist:
        return Response({"detail": "Not found."}, status=404)
    if request.user.role == "branch_user" and report.branch != request.user.branch:
        return Response({"detail": "Forbidden."}, status=403)

    metric    = request.query_params.get("metric")
    start, end = _day_range(report.report_date)
    branch_id  = report.branch_id

    if metric == "student_registrations":
        return Response(list(Student.objects.filter(branch_id=branch_id, created_at__range=[start, end]).values(
            "id", "full_name", "admission_number", "phone", "status", "created_at")))
    if metric == "student_course_registrations":
        course_id = request.query_params.get("course")
        qs = StudentCourse.objects.filter(student__branch_id=branch_id, registration_date__range=[start, end])
        if course_id:
            qs = qs.filter(course_id=int(course_id))
        return Response(list(qs.values(
            "id", "student__full_name", "student__admission_number",
            "course__class_name", "status", "amount_agreed", "registration_date")))
    if metric == "payments":
        return Response(list(PaymentTransaction.objects.filter(
            student__branch_id=branch_id, status="completed", created_at__range=[start, end]
        ).values("id", "student__full_name", "student__admission_number", "amount", "reference_code", "payment_method", "created_at")))
    if metric == "exam_bookings":
        return Response(list(ExamBooking.objects.filter(
            student__branch_id=branch_id, created_at__range=[start, end]
        ).values("id", "student__full_name", "student__admission_number", "exam__exam_name", "exam__exam_date", "status", "created_at")))
    return Response({"detail": "Unknown metric."}, status=400)


# ── Daily summary (live metrics merged into report data) ─────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def daily_summary(request):
    user     = request.user
    date_str = request.query_params.get("date") or timezone.now().date().isoformat()
    try:
        target_date = _parse_date(date_str)
    except ValueError:
        return Response({"detail": "Invalid date."}, status=400)

    start, end = _day_range(target_date)

    def live_metrics_for_branch(branch_id):
        payments = PaymentTransaction.objects.filter(
            student__branch_id=branch_id, status="completed",
            student__isnull=False, created_at__range=[start, end]
        ).aggregate(count=Count("id"), total=Sum("amount"))
        return {
            "payment_total":                float(payments["total"] or 0),
            "payment_count":                payments["count"] or 0,
            "student_registrations":        Student.objects.filter(branch_id=branch_id, created_at__range=[start, end]).count(),
            "student_course_registrations": StudentCourse.objects.filter(student__branch_id=branch_id, registration_date__range=[start, end]).count(),
            "exam_bookings_count":          ExamBooking.objects.filter(student__branch_id=branch_id, created_at__range=[start, end]).count(),
        }

    def merge(report, branch_id):
        if not report:
            return None
        data = ReportSerializer(report).data
        data.update(live_metrics_for_branch(branch_id))
        return data

    if user.role == "branch_user":
        report = Report.objects.filter(branch=user.branch, report_date=target_date).select_related("branch", "created_by").first()
        return Response({"date": date_str, "branches": [{
            "branch_id": user.branch.id, "branch_name": user.branch.name,
            "report": merge(report, user.branch.id),
        }]})

    all_branches = Branch.objects.all().order_by("name")
    reports_map  = {r.branch_id: r for r in Report.objects.filter(report_date=target_date).select_related("branch", "created_by")}
    result = []
    for b in all_branches:
        result.append({
            "branch_id": b.id, "branch_name": b.name,
            "report": merge(reports_map.get(b.id), b.id),
        })
    return Response({"date": date_str, "branches": result})


# ── Live KPI Summary ──────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def analytics_kpi_live(request):
    params    = request.query_params
    user      = request.user
    branch_id = user.branch_id if user.role == "branch_user" else params.get("branch")

    rev_qs = _bf(_apply_dt_filters(PaymentTransaction.objects.filter(status="completed", student__isnull=False), params), "student__branch_id", branch_id)
    revenue = float(rev_qs.aggregate(t=Sum("amount"))["t"] or 0)

    stu_qs  = _bf(_apply_dt_filters(Student.objects.all(), params), "branch_id", branch_id)
    sc_qs   = _bf(_apply_date_filters_sc(StudentCourse.objects.all(), params), "student__branch_id", branch_id)
    exam_qs = _bf(_apply_dt_filters(ExamBooking.objects.all(), params), "student__branch_id", branch_id)
    rpt_qs  = _bf(_apply_date_filters_report(Report.objects.all(), params), "branch_id", branch_id)
    rpt_agg = rpt_qs.aggregate(inq=Sum("inquiries"), cnt=Count("id"))

    return Response({
        "revenue":       revenue,
        "registrations": stu_qs.count(),
        "enrollments":   sc_qs.count(),
        "exam_bookings": exam_qs.count(),
        "inquiries":     rpt_agg["inq"] or 0,
        "report_count":  rpt_agg["cnt"] or 0,
    })


# ── Payment Type Breakdown ────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def analytics_payment_types(request):
    params    = request.query_params
    branch_id = request.user.branch_id if request.user.role == "branch_user" else params.get("branch")
    qs = _bf(_apply_dt_filters(PaymentTransaction.objects.filter(status="completed", student__isnull=False), params), "student__branch_id", branch_id)
    rows = qs.values("payment_type").annotate(count=Count("id"), amount=Sum("amount")).order_by("-amount")
    return Response([{"payment_type": r["payment_type"], "count": r["count"], "amount": float(r["amount"] or 0)} for r in rows])


# ── Branch Comparison ─────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def analytics_branch_comparison(request):
    if request.user.role != "super_admin":
        return Response({"detail": "Forbidden."}, status=403)
    params     = request.query_params
    metric     = params.get("metric", "payment_total")
    branch_ids = [int(b) for b in params.getlist("branches") if b.isdigit()]
    course_id  = params.get("course")

    def fb(qs, field):
        return qs.filter(**{f"{field}__in": branch_ids}) if branch_ids else qs

    if metric == "payment_total":
        qs = fb(_apply_dt_filters(PaymentTransaction.objects.filter(status="completed", student__isnull=False), params), "student__branch_id")
        rows = qs.values("student__branch__name").annotate(value=Sum("amount")).order_by("-value")
        return Response([{"branch": r["student__branch__name"], "value": float(r["value"] or 0)} for r in rows])

    if metric == "student_registrations":
        qs = fb(_apply_dt_filters(Student.objects.all(), params), "branch_id")
        rows = qs.values("branch__name").annotate(value=Count("id")).order_by("-value")
        return Response([{"branch": r["branch__name"], "value": r["value"]} for r in rows])

    if metric == "student_course_registrations":
        qs = fb(_apply_date_filters_sc(StudentCourse.objects.all(), params), "student__branch_id")
        if course_id:
            qs = qs.filter(course_id=int(course_id))
        rows = qs.values("student__branch__name").annotate(value=Count("id")).order_by("-value")
        return Response([{"branch": r["student__branch__name"], "value": r["value"]} for r in rows])

    if metric == "exam_bookings_count":
        qs = fb(_apply_dt_filters(ExamBooking.objects.all(), params), "student__branch_id")
        rows = qs.values("student__branch__name").annotate(value=Count("id")).order_by("-value")
        return Response([{"branch": r["student__branch__name"], "value": r["value"]} for r in rows])

    if metric == "inquiries":
        qs = _apply_date_filters_report(Report.objects.all(), params)
        if branch_ids:
            qs = qs.filter(branch_id__in=branch_ids)
        rows = qs.values("branch__name").annotate(value=Sum(metric)).order_by("-value")
        return Response([{"branch": r["branch__name"], "value": r["value"] or 0} for r in rows])

    return Response({"detail": "Unsupported metric."}, status=400)


# ── Time Series ───────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def analytics_time_series(request):
    user          = request.user
    params        = request.query_params
    metric        = params.get("metric", "payment_total")
    date_from_str = params.get("date_from")
    date_to_str   = params.get("date_to")
    branch_id_raw = params.get("branch")
    branch_id     = int(branch_id_raw) if branch_id_raw and branch_id_raw.isdigit() else None
    course_id_raw = params.get("course")
    course_id     = int(course_id_raw) if course_id_raw and course_id_raw.isdigit() else None

    if user.role == "branch_user":
        branch_id = user.branch_id

    LIVE    = {"payment_total", "student_registrations", "student_course_registrations", "exam_bookings_count"}
    MANUAL  = {"inquiries"}
    if metric not in LIVE | MANUAL:
        return Response({"detail": "Invalid metric."}, status=400)

    if date_from_str and date_to_str:
        try:    gran = _auto_granularity(_parse_date(date_from_str), _parse_date(date_to_str))
        except: gran = "monthly"
    else:
        gran = "monthly"
    TF = {"daily": TruncDay, "weekly": TruncWeek, "monthly": TruncMonth}[gran]

    def period_str(p):
        return (p.date() if hasattr(p, "date") else p).strftime("%Y-%m-%d")

    if metric == "payment_total":
        qs = PaymentTransaction.objects.filter(status="completed", student__isnull=False)
        if branch_id:     qs = qs.filter(student__branch_id=branch_id)
        if date_from_str: qs = qs.filter(created_at__gte=timezone.make_aware(datetime.combine(_parse_date(date_from_str), time.min)))
        if date_to_str:   qs = qs.filter(created_at__lte=timezone.make_aware(datetime.combine(_parse_date(date_to_str), time.max)))
        rows = qs.annotate(period=TF("created_at")).values("period").annotate(value=Sum("amount")).order_by("period")
        return Response({"data": [{"period": period_str(r["period"]), "value": float(r["value"] or 0)} for r in rows], "granularity": gran})

    if metric == "student_registrations":
        qs = Student.objects.all()
        if branch_id:     qs = qs.filter(branch_id=branch_id)
        if date_from_str: qs = qs.filter(created_at__gte=timezone.make_aware(datetime.combine(_parse_date(date_from_str), time.min)))
        if date_to_str:   qs = qs.filter(created_at__lte=timezone.make_aware(datetime.combine(_parse_date(date_to_str), time.max)))
        rows = qs.annotate(period=TF("created_at")).values("period").annotate(value=Count("id")).order_by("period")
        return Response({"data": [{"period": period_str(r["period"]), "value": r["value"]} for r in rows], "granularity": gran})

    if metric == "student_course_registrations":
        qs = StudentCourse.objects.all()
        if branch_id:     qs = qs.filter(student__branch_id=branch_id)
        if course_id:     qs = qs.filter(course_id=course_id)
        if date_from_str: qs = qs.filter(registration_date__gte=date_from_str)
        if date_to_str:   qs = qs.filter(registration_date__lte=date_to_str)
        rows = qs.annotate(period=TF("registration_date")).values("period").annotate(value=Count("id")).order_by("period")
        return Response({"data": [{"period": period_str(r["period"]), "value": r["value"]} for r in rows], "granularity": gran})

    if metric == "exam_bookings_count":
        qs = ExamBooking.objects.all()
        if branch_id:     qs = qs.filter(student__branch_id=branch_id)
        if date_from_str: qs = qs.filter(created_at__gte=timezone.make_aware(datetime.combine(_parse_date(date_from_str), time.min)))
        if date_to_str:   qs = qs.filter(created_at__lte=timezone.make_aware(datetime.combine(_parse_date(date_to_str), time.max)))
        rows = qs.annotate(period=TF("created_at")).values("period").annotate(value=Count("id")).order_by("period")
        return Response({"data": [{"period": period_str(r["period"]), "value": r["value"]} for r in rows], "granularity": gran})

    # inquiries — from Report model
    qs = Report.objects.all()
    if user.role == "branch_user": qs = qs.filter(branch=user.branch)
    elif branch_id:                qs = qs.filter(branch_id=branch_id)
    if date_from_str: qs = qs.filter(report_date__gte=date_from_str)
    if date_to_str:   qs = qs.filter(report_date__lte=date_to_str)
    rows = qs.annotate(period=TF("report_date")).values("period").annotate(value=Sum(metric)).order_by("period")
    return Response({"data": [{"period": r["period"].strftime("%Y-%m-%d"), "value": r["value"] or 0} for r in rows], "granularity": gran})


# ── Export Summary (JSON for frontend preview) ────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def analytics_export_summary(request):
    params    = request.query_params
    branch_id = request.user.branch_id if request.user.role == "branch_user" else params.get("branch")
    sections  = set(params.getlist("sections")) or {"operational", "revenue", "branch_perf", "course_enroll"}
    result    = {}

    if "operational" in sections:
        stu_qs  = _bf(_apply_dt_filters(Student.objects.all(), params), "branch_id", branch_id)
        sc_qs   = _bf(_apply_date_filters_sc(StudentCourse.objects.all(), params), "student__branch_id", branch_id)
        exam_qs = _bf(_apply_dt_filters(ExamBooking.objects.all(), params), "student__branch_id", branch_id)
        rpt_qs  = _bf(_apply_date_filters_report(Report.objects.all(), params), "branch_id", branch_id)
        rpt_agg = rpt_qs.aggregate(inq=Sum("inquiries"), cnt=Count("id"))
        result["operational"] = {
            "registrations":    stu_qs.count(),
            "enrollments":      sc_qs.count(),
            "exam_bookings":    exam_qs.count(),
            "inquiries":        rpt_agg["inq"] or 0,
            "report_count":     rpt_agg["cnt"] or 0,
        }

    if "revenue" in sections:
        rev_qs  = _bf(_apply_dt_filters(PaymentTransaction.objects.filter(status="completed", student__isnull=False), params), "student__branch_id", branch_id)
        by_type = list(rev_qs.values("payment_type").annotate(count=Count("id"), amount=Sum("amount")).order_by("-amount"))
        result["revenue"] = {
            "total":   float(rev_qs.aggregate(t=Sum("amount"))["t"] or 0),
            "by_type": [{"payment_type": r["payment_type"], "count": r["count"], "amount": float(r["amount"] or 0)} for r in by_type],
        }

    if "branch_perf" in sections:
        all_branches = Branch.objects.filter(pk=branch_id) if branch_id else Branch.objects.all()
        rows = []
        for b in all_branches:
            rev  = float(_apply_dt_filters(PaymentTransaction.objects.filter(status="completed", student__branch_id=b.id, student__isnull=False), params).aggregate(t=Sum("amount"))["t"] or 0)
            regs = _apply_dt_filters(Student.objects.filter(branch_id=b.id), params).count()
            enrl = _apply_date_filters_sc(StudentCourse.objects.filter(student__branch_id=b.id), params).count()
            agg  = _apply_date_filters_report(Report.objects.filter(branch_id=b.id), params).aggregate(inq=Sum("inquiries"))
            exms = _apply_dt_filters(ExamBooking.objects.filter(student__branch_id=b.id), params).count()
            rows.append({"branch": b.name, "revenue": rev, "registrations": regs, "enrollments": enrl, "inquiries": agg["inq"] or 0, "exam_bookings": exms})
        rows.sort(key=lambda x: x["revenue"], reverse=True)
        result["branch_perf"] = rows

    if "course_enroll" in sections:
        sc_qs = _bf(_apply_date_filters_sc(StudentCourse.objects.all(), params), "student__branch_id", branch_id)
        rows  = list(sc_qs.values("course__class_name").annotate(count=Count("id")).order_by("-count"))
        result["course_enroll"] = [{"course": r["course__class_name"], "count": r["count"], "revenue": 0} for r in rows]

    return Response(result)


# ── Excel Export (server-side, openpyxl) ─────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_reports_excel(request):
    params    = request.query_params
    branch_id = request.user.branch_id if request.user.role == "branch_user" else params.get("branch")
    sections  = set(params.getlist("sections")) or {"operational", "revenue", "branch_perf"}

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    HDR_FILL = PatternFill("solid", fgColor="111827")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10)

    def make_sheet(title, headers, rows):
        ws = wb.create_sheet(title)
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.fill = HDR_FILL; c.font = HDR_FONT; c.alignment = Alignment(horizontal="center")
        for row in rows:
            ws.append(row)
        for col in ws.columns:
            ws.column_dimensions[col[0].column_letter].width = max((len(str(c.value or "")) for c in col), default=10) + 4
        return ws

    if "operational" in sections:
        stu_qs  = _bf(_apply_dt_filters(Student.objects.all(), params), "branch_id", branch_id)
        sc_qs   = _bf(_apply_date_filters_sc(StudentCourse.objects.all(), params), "student__branch_id", branch_id)
        exam_qs = _bf(_apply_dt_filters(ExamBooking.objects.all(), params), "student__branch_id", branch_id)
        rpt_qs  = _bf(_apply_date_filters_report(Report.objects.all(), params), "branch_id", branch_id)
        rpt_agg = rpt_qs.aggregate(inq=Sum("inquiries"), cnt=Count("id"))
        make_sheet("Operational", ["Metric", "Value"], [
            ["Registrations",       stu_qs.count()],
            ["Enrollments",         sc_qs.count()],
            ["Exam Bookings",       exam_qs.count()],
            ["Inquiries",           rpt_agg["inq"] or 0],
            ["Reports Submitted",   rpt_agg["cnt"] or 0],
        ])

    if "revenue" in sections:
        rev_qs  = _bf(_apply_dt_filters(PaymentTransaction.objects.filter(status="completed", student__isnull=False), params), "student__branch_id", branch_id)
        total   = float(rev_qs.aggregate(t=Sum("amount"))["t"] or 0)
        by_type = list(rev_qs.values("payment_type").annotate(count=Count("id"), amount=Sum("amount")).order_by("-amount"))
        make_sheet("Revenue", ["Payment Type", "Count", "Amount (Ksh)"],
            [[r["payment_type"], r["count"], float(r["amount"] or 0)] for r in by_type] + [["TOTAL", "", total]])

    if "branch_perf" in sections:
        all_branches = Branch.objects.filter(pk=branch_id) if branch_id else Branch.objects.all()
        rows = []
        for b in all_branches:
            rev  = float(_apply_dt_filters(PaymentTransaction.objects.filter(status="completed", student__branch_id=b.id, student__isnull=False), params).aggregate(t=Sum("amount"))["t"] or 0)
            regs = _apply_dt_filters(Student.objects.filter(branch_id=b.id), params).count()
            enrl = _apply_date_filters_sc(StudentCourse.objects.filter(student__branch_id=b.id), params).count()
            agg  = _apply_date_filters_report(Report.objects.filter(branch_id=b.id), params).aggregate(inq=Sum("inquiries"))
            exms = _apply_dt_filters(ExamBooking.objects.filter(student__branch_id=b.id), params).count()
            rows.append([b.name, rev, regs, enrl, agg["inq"] or 0, exms])
        rows.sort(key=lambda x: x[1], reverse=True)
        make_sheet("Branch Performance",
            ["Branch", "Revenue (Ksh)", "Registrations", "Enrollments", "Inquiries", "Exam Bookings"], rows)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    resp = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    resp["Content-Disposition"] = 'attachment; filename="reports_export.xlsx"'
    return resp


# ── Reference data ────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def courses_list(request):
    return Response(list(Course.objects.values("id", "class_name", "category")))

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def branches_list_for_reports(request):
    if request.user.role != "super_admin":
        return Response({"detail": "Forbidden."}, status=403)
    return Response(list(Branch.objects.values("id", "name")))