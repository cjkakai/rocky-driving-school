import re
from django.db import IntegrityError, transaction
from rest_framework import viewsets
from rest_framework.decorators import action, api_view, permission_classes
from rest_framework.exceptions import PermissionDenied, ValidationError
from rest_framework.pagination import PageNumberPagination
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.db.models import Q, Max, Count, Sum, F
from django.utils import timezone
from datetime import timedelta
from .models import Student
from .serializers import StudentSerializer
from students.lifecycle import check_and_apply_pdl_expiry


class StandardPagination(PageNumberPagination):
    page_size = 20
    page_size_query_param = "page_size"
    max_page_size = 100


# Single source of truth for prefetches — used in both _base_queryset and exam_roster
_STUDENT_PREFETCHES = [
    "student_courses__course",
    "student_courses__payments",
    "student_courses__exam_bookings__exam",
    "student_courses__exam_bookings__result",
    "student_courses__pdl_bookings",
    "pdl_bookings",
]


def generate_admission_number(branch):
    """Must be called from within an active transaction.atomic() block."""
    code = branch.branch_code.upper() if branch else "GEN"
    all_nums = (
        Student.objects.select_for_update()
        .filter(branch=branch, admission_number__startswith=code)
        .values_list("admission_number", flat=True)
    )
    max_num = max(
        (int(m.group(0)) for adm in all_nums if (m := re.search(r'\d+$', adm))),
        default=0,
    )
    return f"{code}{str(max_num + 1).zfill(3)}"


class StudentViewSet(viewsets.ModelViewSet):
    serializer_class = StudentSerializer
    permission_classes = [IsAuthenticated]
    pagination_class = StandardPagination

    def perform_create(self, serializer):
        branch = serializer.validated_data.get("branch")
        for attempt in range(3):
            try:
                with transaction.atomic():
                    admission_number = generate_admission_number(branch)
                    serializer.save(admission_number=admission_number, status="active")
                return
            except IntegrityError as e:
                if "admission_number" not in str(e) or attempt == 2:
                    raise

    def perform_update(self, serializer):
        if self.request.user.role != "super_admin":
            raise PermissionDenied("Only super admins can edit students.")
        # Prevent admission_number from being changed via update
        serializer.save()

    def destroy(self, request, *args, **kwargs):
        if request.user.role != "super_admin":
            raise PermissionDenied("Only super admins can delete students.")
        student = self.get_object()
        if student.payments.filter(status="completed").exists():
            raise ValidationError(
                "Cannot delete a student with completed payment records. Offload the student instead."
            )
        return super().destroy(request, *args, **kwargs)

    def _base_queryset(self, user):
        qs = Student.objects.select_related("branch").prefetch_related(*_STUDENT_PREFETCHES)
        if user.role == "branch_user":
            return qs.filter(branch=user.branch) if user.branch else Student.objects.none()
        return qs

    def _apply_filters(self, queryset, params, user):
        needs_distinct = False

        # Student aggregate status
        student_status = params.get("status")
        if student_status and student_status != "all":
            queryset = queryset.filter(status=student_status)
        elif not student_status:
            queryset = queryset.exclude(status="offloaded")

        # Branch (super_admin only)
        branch_id = params.get("branch_id")
        if branch_id and user.role == "super_admin":
            queryset = queryset.filter(branch_id=branch_id)

        # Course filter
        course_id = params.get("course_id")
        if course_id:
            queryset = queryset.filter(student_courses__course_id=course_id)
            needs_distinct = True

        # StudentCourse lifecycle status
        course_status = params.get("course_status")
        if course_status:
            queryset = queryset.filter(student_courses__status=course_status)
            needs_distinct = True

        # Period filter — based on StudentCourse.registration_date
        date_from = params.get("date_from")
        date_to = params.get("date_to")
        if date_from:
            queryset = queryset.filter(student_courses__registration_date__date__gte=date_from)
            needs_distinct = True
        if date_to:
            queryset = queryset.filter(student_courses__registration_date__date__lte=date_to)
            needs_distinct = True

        # PDL expiry filter
        pdl_days = params.get("pdl_days")
        if pdl_days:
            try:
                days = int(pdl_days)
                cutoff = timezone.now() + timedelta(days=days)
                queryset = queryset.filter(
                    pdl_bookings__status="approved",
                    pdl_bookings__approved_at__isnull=False,
                    pdl_bookings__approved_at__gt=timezone.now() - timedelta(days=90),
                    pdl_bookings__approved_at__lte=cutoff - timedelta(days=90),
                )
                needs_distinct = True
            except (ValueError, TypeError):
                pass

        # Exam filter — students who have a booking for the given exam
        exam_id = params.get("exam_id")
        if exam_id:
            queryset = queryset.filter(student_courses__exam_bookings__exam_id=exam_id)
            needs_distinct = True

        # Search
        search = params.get("search")
        if search:
            queryset = queryset.filter(
                Q(full_name__icontains=search)
                | Q(phone__icontains=search)
                | Q(admission_number__icontains=search)
                | Q(id_number__icontains=search)
            )

        return queryset.distinct() if needs_distinct else queryset

    def get_queryset(self):
        user = self.request.user
        params = self.request.query_params
        queryset = self._base_queryset(user)
        queryset = self._apply_filters(queryset, params, user)
        return queryset.annotate(
            latest_course_date=Max("student_courses__registration_date")
        ).order_by("-latest_course_date", "-created_at").distinct()

    def retrieve(self, request, *args, **kwargs):
        instance = self.get_object()
        from students.lifecycle import ACTIVE_STATUSES
        for sc in instance.student_courses.filter(
            status__in=list(ACTIVE_STATUSES)
        ).prefetch_related("pdl_bookings"):
            check_and_apply_pdl_expiry(sc)
        instance.refresh_from_db()
        return Response(self.get_serializer(instance).data)

    @action(detail=False, methods=["get"])
    def exam_roster(self, request):
        """
        Students enrolled in a specific exam, bypassing lifecycle filters.
        Used for bulk SMS targeting — includes offloaded/completed students.
        """
        exam_id = request.query_params.get("exam_id")
        if not exam_id:
            return Response({"detail": "exam_id is required."}, status=400)

        qs = Student.objects.select_related("branch").prefetch_related(
            *_STUDENT_PREFETCHES
        ).filter(student_courses__exam_bookings__exam_id=exam_id)

        if request.user.role == "branch_user" and request.user.branch:
            qs = qs.filter(branch=request.user.branch)
        elif request.user.role == "super_admin":
            branch_id = request.query_params.get("branch_id")
            if branch_id:
                qs = qs.filter(branch_id=branch_id)

        exam_result = request.query_params.get("exam_result")
        if exam_result and exam_result != "all":
            qs = qs.filter(
                student_courses__exam_bookings__result__result=exam_result.lower()
            )

        search = request.query_params.get("search")
        if search:
            qs = qs.filter(
                Q(full_name__icontains=search)
                | Q(phone__icontains=search)
                | Q(admission_number__icontains=search)
                | Q(id_number__icontains=search)
            )

        qs = qs.distinct()
        page = self.paginate_queryset(qs)
        serializer = self.get_serializer(page, many=True)
        return self.get_paginated_response(serializer.data)

    @action(detail=False, methods=["get"])
    def export(self, request):
        """Export students to Excel respecting all active filters."""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
        except ImportError:
            return Response({"detail": "openpyxl is required for export."}, status=500)

        from django.http import HttpResponse

        user = request.user
        params = request.query_params
        queryset = self._base_queryset(user)
        queryset = self._apply_filters(queryset, params, user)
        queryset = queryset.annotate(
            latest_course_date=Max("student_courses__registration_date")
        ).order_by("-latest_course_date", "-created_at")

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Students"

        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill("solid", fgColor="0F172A")
        center = Alignment(horizontal="center")

        headers = [
            "Adm. No", "Full Name", "Phone", "ID Number", "Branch",
            "Student Status", "Course", "Course Status", "Enrollment Date",
            "Amount Agreed", "Total Paid", "Balance",
        ]
        ws.append(headers)
        for col, _ in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center

        # When an exam filter is active, restrict each student's rows to only
        # the course(s) linked to that exam — not all their courses.
        exam_id = params.get("exam_id")

        for student in queryset:
            all_courses = list(student.student_courses.all())
            if exam_id:
                courses = [
                    sc for sc in all_courses
                    if any(eb.exam_id == int(exam_id) for eb in sc.exam_bookings.all())
                ]
            else:
                courses = [sc for sc in all_courses if sc.status != "transferred"]

            if not courses:
                ws.append([
                    student.admission_number, student.full_name, student.phone,
                    student.id_number, student.branch.name if student.branch else "",
                    student.status, "", "", "", "", "", "",
                ])
            else:
                for sc in courses:
                    agreed = float(sc.amount_agreed or 0)
                    paid = sum(float(p.amount) for p in sc.payments.filter(status="completed"))
                    ws.append([
                        student.admission_number,
                        student.full_name,
                        student.phone,
                        student.id_number,
                        student.branch.name if student.branch else "",
                        student.status,
                        sc.course.class_name if sc.course else "",
                        sc.status,
                        sc.registration_date.strftime("%d/%m/%Y") if sc.registration_date else "",
                        agreed,
                        paid,
                        round(agreed - paid, 2),
                    ])

        # Auto-size columns
        for col in ws.columns:
            max_len = max((len(str(cell.value or "")) for cell in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = 'attachment; filename="students_export.xlsx"'
        wb.save(response)
        return response


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def students_summary(request):
    """Aggregated student/course stats for the current filters — no pagination."""
    viewset = StudentViewSet()
    viewset.request = request
    qs = viewset._base_queryset(request.user)
    qs = viewset._apply_filters(qs, request.query_params, request.user)

    student_count = qs.count()

    params = request.query_params
    from academics.models import StudentCourse

    # Start from courses belonging to the matched students, excluding transferred
    sc_qs = StudentCourse.objects.filter(student__in=qs).exclude(status="transferred")

    # Mirror course-level filters onto sc_qs so financial totals reflect
    # only the courses that caused the student to appear — not all their courses.
    course_status = params.get("course_status")
    if course_status:
        sc_qs = sc_qs.filter(status=course_status)

    course_id = params.get("course_id")
    if course_id:
        sc_qs = sc_qs.filter(course_id=course_id)

    date_from = params.get("date_from")
    if date_from:
        sc_qs = sc_qs.filter(registration_date__date__gte=date_from)

    date_to = params.get("date_to")
    if date_to:
        sc_qs = sc_qs.filter(registration_date__date__lte=date_to)

    exam_id = params.get("exam_id")
    if exam_id:
        sc_qs = sc_qs.filter(exam_bookings__exam_id=exam_id)

    # Resolve to distinct course IDs to prevent fan-out from exam_bookings join
    sc_ids = sc_qs.values_list("id", flat=True).distinct()
    sc_qs = StudentCourse.objects.filter(id__in=sc_ids)

    course_counts = dict(
        sc_qs.values("status").annotate(n=Count("id")).values_list("status", "n")
    )
    total_courses = sum(course_counts.values())

    pending_payment = (
        sc_qs
        .annotate(total_paid=Sum("payments__amount", filter=Q(payments__status="completed")))
        .filter(amount_agreed__gt=0)
        .filter(Q(total_paid__isnull=True) | Q(total_paid__lt=F("amount_agreed")))
        .values("student_id")
        .distinct()
        .count()
    )

    completed_courses = course_counts.get("completed", 0)
    failed_retake = course_counts.get("failed", 0) + course_counts.get("retake_booked", 0)

    # Two separate aggregates — combining them in one call causes Django to JOIN
    # both payments and amount_agreed simultaneously, multiplying amount_agreed
    # by the number of payment rows (fan-out). Split calls avoid this entirely.
    total_agreed = float(sc_qs.aggregate(v=Sum("amount_agreed"))["v"] or 0)
    total_paid   = float(sc_qs.aggregate(v=Sum("payments__amount", filter=Q(payments__status="completed")))["v"] or 0)

    return Response({
        "student_count": student_count,
        "total_courses": total_courses,
        "completed_courses": completed_courses,
        "failed_retake": failed_retake,
        "pending_payment": pending_payment,
        "course_counts": course_counts,
        "total_agreed": total_agreed,
        "total_paid": total_paid,
        "total_balance": round(total_agreed - total_paid, 2),
    })
