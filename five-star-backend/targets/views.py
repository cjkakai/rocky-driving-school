import io
from datetime import date, timedelta

from django.db.models import Sum, Count
from django.http import HttpResponse
from django.utils import timezone
from rest_framework.decorators import api_view, permission_classes
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import RevenueTarget, RegistrationTarget
from .serializers import RevenueTargetSerializer, RegistrationTargetSerializer
from finance.models import PaymentTransaction
from academics.models import StudentCourse
from branches.models import Branch


# ── Helpers ───────────────────────────────────────────────────────────────────

def _current_week():
    today = timezone.now().date()
    iso = today.isocalendar()
    return iso[0], iso[1]


def _current_month():
    today = timezone.now().date()
    return today.year, today.month


def _week_date_range(year, week):
    jan4 = date(year, 1, 4)
    monday = jan4 + timedelta(weeks=week - 1, days=-jan4.weekday())
    return monday, monday + timedelta(days=6)


def _achievement_status(pct):
    if pct > 100:
        return "Above Target"
    if pct == 100:
        return "On Target"
    return "Below Target"


def _resolve_branch(request):
    if request.user.role != "super_admin":
        return request.user.branch_id
    param = request.query_params.get("branch", "")
    try:
        val = int(param)
        return val if val > 0 else None
    except (ValueError, TypeError):
        return None


def _parse_int(value, default):
    """Safely parse a query param integer, rejecting non-numeric strings."""
    try:
        val = int(value)
        return val if val > 0 else default
    except (ValueError, TypeError):
        return default


def _make_excel_response(wb, filename):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    res = HttpResponse(buf.read(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    res["Content-Disposition"] = f'attachment; filename="{filename}"'
    return res


def _style_header(ws, headers):
    fill = PatternFill("solid", fgColor="111827")
    font = Font(bold=True, color="FFFFFF", size=10)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center")


def _autofit(ws):
    for col in ws.columns:
        width = max(len(str(c.value or "")) for c in col) + 4
        ws.column_dimensions[col[0].column_letter].width = width


# ── Current Period (server time) ─────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def current_period(request):
    year, week = _current_week()
    year_m, month = _current_month()
    return Response({"year": year, "week": week, "month": month})


# ── Revenue Targets CRUD (admin only) ─────────────────────────────────────────

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def revenue_target_list_create(request):
    if request.user.role != "super_admin":
        return Response({"detail": "Forbidden."}, status=403)

    if request.method == "GET":
        year, week = _current_week()
        qs = RevenueTarget.objects.filter(year=year, week=week).select_related("branch")
        return Response(RevenueTargetSerializer(qs, many=True).data)

    data = request.data
    year = data.get("year") or _current_week()[0]
    week = data.get("week") or _current_week()[1]
    serializer = RevenueTargetSerializer(data={**data, "year": year, "week": week})
    serializer.is_valid(raise_exception=True)
    obj, _ = RevenueTarget.objects.update_or_create(
        branch_id=serializer.validated_data["branch"].id,
        year=serializer.validated_data["year"],
        week=serializer.validated_data["week"],
        defaults={"target_amount": serializer.validated_data["target_amount"], "created_by": request.user},
    )
    return Response(RevenueTargetSerializer(obj).data, status=status.HTTP_200_OK)


# ── Registration Targets CRUD (admin only) ────────────────────────────────────

@api_view(["GET", "POST"])
@permission_classes([IsAuthenticated])
def registration_target_list_create(request):
    if request.user.role != "super_admin":
        return Response({"detail": "Forbidden."}, status=403)

    if request.method == "GET":
        year, month = _current_month()
        qs = RegistrationTarget.objects.filter(year=year, month=month).select_related("branch")
        return Response(RegistrationTargetSerializer(qs, many=True).data)

    data = request.data
    year = data.get("year") or _current_month()[0]
    month = data.get("month") or _current_month()[1]
    serializer = RegistrationTargetSerializer(data={**data, "year": year, "month": month})
    serializer.is_valid(raise_exception=True)
    obj, _ = RegistrationTarget.objects.update_or_create(
        branch_id=serializer.validated_data["branch"].id,
        year=serializer.validated_data["year"],
        month=serializer.validated_data["month"],
        defaults={"target_count": serializer.validated_data["target_count"], "created_by": request.user},
    )
    return Response(RegistrationTargetSerializer(obj).data, status=status.HTTP_200_OK)


# ── Revenue KPI ───────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def revenue_kpi(request):
    branch_id = _resolve_branch(request)
    p = request.query_params
    year = _parse_int(p.get("year"), _current_week()[0])
    week = _parse_int(p.get("week"), _current_week()[1])
    monday, sunday = _week_date_range(year, week)

    target_qs = RevenueTarget.objects.filter(year=year, week=week)
    if branch_id:
        target_qs = target_qs.filter(branch_id=branch_id)
    total_target = float(target_qs.aggregate(t=Sum("target_amount"))["t"] or 0)

    achieved_qs = PaymentTransaction.objects.filter(
        status="completed", student__isnull=False,
        created_at__date__gte=monday, created_at__date__lte=sunday,
    )
    if branch_id:
        achieved_qs = achieved_qs.filter(student__branch_id=branch_id)
    total_achieved = float(achieved_qs.aggregate(t=Sum("amount"))["t"] or 0)

    pct = round((total_achieved / total_target * 100), 1) if total_target else 0
    return Response({
        "year": year, "week": week,
        "week_label": f"W{week} {year} ({monday.strftime('%d %b')}–{sunday.strftime('%d %b')})",
        "total_target": total_target, "total_achieved": total_achieved,
        "pct": pct, "status": _achievement_status(pct),
    })


# ── Registration KPI ──────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def registration_kpi(request):
    branch_id = _resolve_branch(request)
    p = request.query_params
    year  = _parse_int(p.get("year"),  _current_month()[0])
    month = _parse_int(p.get("month"), _current_month()[1])

    target_qs = RegistrationTarget.objects.filter(year=year, month=month)
    if branch_id:
        target_qs = target_qs.filter(branch_id=branch_id)
    total_target = int(target_qs.aggregate(t=Sum("target_count"))["t"] or 0)

    achieved_qs = StudentCourse.objects.filter(registration_date__year=year, registration_date__month=month)
    if branch_id:
        achieved_qs = achieved_qs.filter(student__branch_id=branch_id)
    total_achieved = achieved_qs.count()

    pct = round((total_achieved / total_target * 100), 1) if total_target else 0
    return Response({
        "year": year, "month": month,
        "month_label": date(year, month, 1).strftime("%B %Y"),
        "total_target": total_target, "total_achieved": total_achieved,
        "pct": pct, "status": _achievement_status(pct),
    })


# ── Revenue Branch Breakdown ──────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def revenue_branches(request):
    branch_id = _resolve_branch(request)
    p = request.query_params
    year = _parse_int(p.get("year"), _current_week()[0])
    week = _parse_int(p.get("week"), _current_week()[1])
    monday, sunday = _week_date_range(year, week)

    target_qs = RevenueTarget.objects.filter(year=year, week=week).select_related("branch")
    if branch_id:
        target_qs = target_qs.filter(branch_id=branch_id)
    targets = {t.branch_id: (t.branch.name, float(t.target_amount)) for t in target_qs}

    achieved_qs = PaymentTransaction.objects.filter(
        status="completed", student__isnull=False,
        created_at__date__gte=monday, created_at__date__lte=sunday,
    )
    if branch_id:
        achieved_qs = achieved_qs.filter(student__branch_id=branch_id)
    achieved = {
        r["student__branch_id"]: (r["student__branch__name"], float(r["achieved"] or 0))
        for r in achieved_qs.values("student__branch_id", "student__branch__name").annotate(achieved=Sum("amount"))
    }

    result = []
    for bid in set(targets):  # only branches with a target set
        name = targets[bid][0]
        t = targets[bid][1]
        a = achieved.get(bid, (None, 0))[1]
        pct = round((a / t * 100), 1) if t else 0
        result.append({"branch_id": bid, "branch": name, "target": t, "achieved": a, "pct": pct, "status": _achievement_status(pct)})

    result.sort(key=lambda x: x["pct"], reverse=True)
    return Response(result)


# ── Registration Branch Breakdown ─────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def registration_branches(request):
    branch_id = _resolve_branch(request)
    p = request.query_params
    year  = _parse_int(p.get("year"),  _current_month()[0])
    month = _parse_int(p.get("month"), _current_month()[1])

    target_qs = RegistrationTarget.objects.filter(year=year, month=month).select_related("branch")
    if branch_id:
        target_qs = target_qs.filter(branch_id=branch_id)
    targets = {t.branch_id: (t.branch.name, int(t.target_count)) for t in target_qs}

    achieved_qs = StudentCourse.objects.filter(registration_date__year=year, registration_date__month=month)
    if branch_id:
        achieved_qs = achieved_qs.filter(student__branch_id=branch_id)
    achieved = {
        r["student__branch_id"]: (r["student__branch__name"], r["achieved"])
        for r in achieved_qs.values("student__branch_id", "student__branch__name").annotate(achieved=Count("id"))
    }

    result = []
    for bid in set(targets):  # only branches with a target set
        name = targets[bid][0]
        t = targets[bid][1]
        a = achieved.get(bid, (None, 0))[1]
        pct = round((a / t * 100), 1) if t else 0
        result.append({"branch_id": bid, "branch": name, "target": t, "achieved": a, "pct": pct, "status": _achievement_status(pct)})

    result.sort(key=lambda x: x["pct"], reverse=True)
    return Response(result)


# ── Revenue Trend ─────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def revenue_trend(request):
    branch_id = _resolve_branch(request)
    year, current_week = _current_week()
    rows = []
    for offset in range(7, -1, -1):
        w = current_week - offset
        y = year
        if w <= 0:
            y -= 1
            w += 52
        monday, sunday = _week_date_range(y, w)

        target_qs = RevenueTarget.objects.filter(year=y, week=w)
        if branch_id:
            target_qs = target_qs.filter(branch_id=branch_id)
        target = float(target_qs.aggregate(t=Sum("target_amount"))["t"] or 0)

        achieved_qs = PaymentTransaction.objects.filter(
            status="completed", student__isnull=False,
            created_at__date__gte=monday, created_at__date__lte=sunday,
        )
        if branch_id:
            achieved_qs = achieved_qs.filter(student__branch_id=branch_id)
        achieved = float(achieved_qs.aggregate(t=Sum("amount"))["t"] or 0)

        pct = round((achieved / target * 100), 1) if target else 0
        rows.append({"label": f"W{w}", "target": target, "achieved": achieved, "pct": pct})

    return Response(rows)


# ── Registration Trend ────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def registration_trend(request):
    branch_id = _resolve_branch(request)
    today = timezone.now().date()
    rows = []
    for offset in range(5, -1, -1):
        m = today.month - offset
        y = today.year
        if m <= 0:
            y -= 1
            m += 12

        target_qs = RegistrationTarget.objects.filter(year=y, month=m)
        if branch_id:
            target_qs = target_qs.filter(branch_id=branch_id)
        target = target_qs.aggregate(t=Sum("target_count"))["t"] or 0

        achieved_qs = StudentCourse.objects.filter(registration_date__year=y, registration_date__month=m)
        if branch_id:
            achieved_qs = achieved_qs.filter(student__branch_id=branch_id)
        achieved = achieved_qs.count()

        pct = round((achieved / target * 100), 1) if target else 0
        rows.append({"label": date(y, m, 1).strftime("%b %Y"), "target": target, "achieved": achieved, "pct": pct})

    return Response(rows)


# ── Combined Summary (shared logic) ──────────────────────────────────────────

def _build_summary_rows(branch_id, params):
    year_w = _parse_int(params.get("year"),  _current_week()[0])
    week   = _parse_int(params.get("week"),   _current_week()[1])
    year_m = _parse_int(params.get("year"),   _current_month()[0])
    month  = _parse_int(params.get("month"),  _current_month()[1])
    monday, sunday = _week_date_range(year_w, week)

    rev_target_qs = RevenueTarget.objects.filter(year=year_w, week=week).select_related("branch")
    if branch_id:
        rev_target_qs = rev_target_qs.filter(branch_id=branch_id)
    rev_targets = {t.branch_id: (t.branch.name, float(t.target_amount)) for t in rev_target_qs}

    rev_achieved_qs = PaymentTransaction.objects.filter(
        status="completed", student__isnull=False,
        created_at__date__gte=monday, created_at__date__lte=sunday,
    )
    if branch_id:
        rev_achieved_qs = rev_achieved_qs.filter(student__branch_id=branch_id)
    rev_achieved = {
        r["student__branch_id"]: float(r["achieved"] or 0)
        for r in rev_achieved_qs.values("student__branch_id").annotate(achieved=Sum("amount"))
    }

    reg_target_qs = RegistrationTarget.objects.filter(year=year_m, month=month).select_related("branch")
    if branch_id:
        reg_target_qs = reg_target_qs.filter(branch_id=branch_id)
    reg_targets = {t.branch_id: (t.branch.name, int(t.target_count)) for t in reg_target_qs}

    reg_achieved_qs = StudentCourse.objects.filter(registration_date__year=year_m, registration_date__month=month)
    if branch_id:
        reg_achieved_qs = reg_achieved_qs.filter(student__branch_id=branch_id)
    reg_achieved = {
        r["student__branch_id"]: r["achieved"]
        for r in reg_achieved_qs.values("student__branch_id").annotate(achieved=Count("id"))
    }

    rows = []
    for bid in set(rev_targets):
        name = rev_targets[bid][0]
        t = rev_targets[bid][1]
        a = rev_achieved.get(bid, 0)
        pct = round((a / t * 100), 1) if t else 0
        rows.append({"branch": name, "metric": "Revenue", "target": t, "achieved": a, "diff": a - t, "pct": pct, "status": _achievement_status(pct)})

    for bid in set(reg_targets):
        name = reg_targets[bid][0]
        t = reg_targets[bid][1]
        a = reg_achieved.get(bid, 0)
        pct = round((a / t * 100), 1) if t else 0
        rows.append({"branch": name, "metric": "Registrations", "target": t, "achieved": a, "diff": a - t, "pct": pct, "status": _achievement_status(pct)})

    rows.sort(key=lambda x: (x["branch"], x["metric"]))
    return rows


@api_view(["GET"])
@permission_classes([IsAuthenticated])
def summary(request):
    rows = _build_summary_rows(_resolve_branch(request), request.query_params)
    return Response(rows)


# ── Export ────────────────────────────────────────────────────────────────────

@api_view(["GET"])
@permission_classes([IsAuthenticated])
def export_summary(request):
    rows = _build_summary_rows(_resolve_branch(request), request.query_params)

    metric = request.query_params.get("metric", "")
    if metric:
        rows = [r for r in rows if r["metric"] == metric]

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Targets"
    headers = ["Branch", "Metric", "Target", "Achieved", "Difference", "Achievement %", "Status"]
    _style_header(ws, headers)
    for r in rows:
        ws.append([r["branch"], r["metric"], r["target"], r["achieved"], r["diff"], r["pct"], r["status"]])
    _autofit(ws)
    return _make_excel_response(wb, "targets_summary.xlsx")
